#author_elvis
#-*- coding:utf-8 -*-
import os
import urllib
import urllib.parse
import urllib.request
import urllib.response
import time
import requests
import openpyxl
import xlsxwriter
from bs4 import BeautifulSoup
from tkinter import Tk
from tkinter import Frame
from tkinter import Label
from tkinter import Button
from tkinter import RIDGE
from tkinter import BOTH
from tkinter import RIGHT
from tkinter import LEFT
import webbrowser
import chardet
import re

header = {
    'User-Agent':'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36',
}
docexcel = 'CatEyes.xlsx'
urllist = []
keylist = []
groups = []

def load():
    if os.path.exists(docexcel) == True:
        openpath = openpyxl.load_workbook(docexcel)
        sheet = openpath.get_sheet_names()
        ws = openpath.get_sheet_by_name(sheet[0])
        rx = ws.max_row
        for i in range(0, int(rx)-1):
            eachurl = ws.cell(row=i + 2, column=1).value
            if not eachurl == None:
                urllist.append(eachurl)
            eachkey = ws.cell(row=i + 2, column=2).value
            if not eachkey == None:
                keylist.append(eachkey)
        openpath.save(docexcel)
    else:
        workbook = xlsxwriter.Workbook(docexcel)
        worksheet = workbook.add_worksheet()
        worksheet.write('A1', 'Monitoring websites')
        worksheet.write('B1', 'Monitoring keys')
        workbook.close()

def buildurl():
    num = 1
    enble = True
    while enble == True:
        url = input('Please input the websites you want to monitor ' + str(num) + '(If none, press Enter to continue):')
        if url == '':
            break
        if not re.search("http", url):
            url = 'http://' + url + '/'
        print(url)
        urllist.append(url)
        num += 1
        add = input('If you want to append other websites?(Y/N)')
        if add == 'Y':
            enble = True
        else:
            wb = openpyxl.load_workbook(docexcel)
            sheet = wb.get_sheet_names()
            ws = wb.get_sheet_by_name(sheet[0])
            num = len(urllist)
            for i in range(0, int(num)):
                ws.cell(row=i + 2, column=1, value=urllist[i])
            wb.save(docexcel)
            enble = False

def buildkey():
    num = 1
    enble = True
    while enble == True:
        key = input('Please input the keys you want to monitor ' + str(num) + '(If none, press Enter to continue):')
        if key == '':
            break
        keylist.append(key)
        num += 1
        add = input('If you want to append other keys?(Y/N)')
        if add == 'Y':
            enble = True
        else:
            wb = openpyxl.load_workbook(docexcel)
            sheet = wb.get_sheet_names()
            ws = wb.get_sheet_by_name(sheet[0])
            num = len(keylist)
            for i in range(0, int(num)):
                ws.cell(row=i + 2, column=2, value=keylist[i])
            wb.save(docexcel)
            enble = False

def matchkey():
    num1 = len(urllist)
    num2 = len(keylist)
    for i in range(0, num1):
        num = []
        url = urllist[i].strip()
        print(url)
        data = urllib.request.urlopen(url).read()
        chardetx = chardet.detect(data)
        html = requests.get(url)
        html.encoding = chardetx['encoding']
        html = html.text
        content = BeautifulSoup(html, 'lxml')
        for i in range(0,num2):
            key = str(keylist[i])
            pattern = re.compile(key, re.S)
            info = content.find_all('a')
            count = 0
            for each in info:
                if re.search(pattern, str(each)):
                    num.append(count)
                count += 1
            for i in num:
                group = []
                link = info[i].get('href')
                group.append(link)
                title = info[i].get_text().strip()
                group.append(title)
                if group not in groups:
                    print(groups)
                    if re.search(pattern, str(group[1])):
                        print(re.search(pattern, str(group[1])))
                        groups.append(group)

def skippage(linkinfo, root):
    webbrowser.open(linkinfo)
    root.destroy()

def showMessage(group):
    linkinfo = group[0]
    lentitle = len(group[1])
    if lentitle > 20 and lentitle <= 40:
        titleinfo = group[1][:20] + '\n' + group[1][20:]
    elif lentitle > 40:
        titleinfo = group[1][:20] + '\n' + group[1][20:40] + '\n' + group[1][40:]
    else:
        titleinfo = group[1]
    print(titleinfo)
    root = Tk()
    screenwidth = root.winfo_screenwidth()
    screenheight = root.winfo_screenheight()
    root.resizable(False, False)
    root.title("Hot News")
    frame = Frame(root, relief=RIDGE, borderwidth=3)
    frame.pack(fill=BOTH, expand=1)
    label = Label(frame, text=titleinfo, font="Corsiva -13")
    label.pack(fill=BOTH, expand=1)
    button1 = Button(frame, text="Check", font="Cooper -15 bold", command=lambda :skippage(linkinfo, root))
    button2 = Button(frame, text="Known", font="Cooper -15 bold", command=root.destroy)
    button1.pack(side=LEFT)
    button2.pack(side=RIGHT)
    root.update_idletasks()
    root.deiconify()
    root.geometry('300x100+%s+%s' % (screenwidth - 350, screenheight - 200))
    root.deiconify()
    root.mainloop()

if __name__ == '__main__':
    load()
    judge = input('If you want to build new items?(Y/N)')
    if judge == 'Y':
        buildurl()
        buildkey()
    print('Beginning...')
    while True:
        q = len(groups)
        print(q)
        matchkey()
        qq = len(groups)
        print(qq)
        if qq != q:
            for i in range(int(q), int(qq)):
                group = groups[i]
                showMessage(group)
        time.sleep(2)