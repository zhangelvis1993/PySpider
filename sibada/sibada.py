#_*_coding:utf-8_*_
import os
import urllib
import urllib.parse
import urllib.request
import urllib.response
from urllib.parse import quote
import time
from selenium import webdriver
import pytesseract
import requests
from lxml import etree
import openpyxl
import xlsxwriter
from bs4 import BeautifulSoup
from tkinter import *
import chardet
from tkinter import ttk
import http.client
import queue
import urllib3

header = {
    'User-Agent':'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36',
}
doctxt = '斯巴达.xlsx'
#项目名和选择方式
infovalue1 = ['', '', '']
#登录/翻页/详情设置/抓取设置的append
infovalue2 = [['','','','','','','','','','','','','','','','']]
#读取已有设置进行自动化
infovalue3 = []

#工具
    #获取网络时间
def get_webservertime():
    host = 'www.baidu.com'
    conn=http.client.HTTPConnection(host)
    conn.request("GET", "/")
    r = conn.getresponse()
    # r.getheaders() #获取所有的http头
    ts = r.getheader('date')  # 获取http头date部分
    print(ts)
    # 将GMT时间转换成北京时间
    ltime = time.strptime(ts[5:25], "%d %b %Y %H:%M:%S")
    ttime = time.localtime(time.mktime(ltime) + 8 * 60 * 60)
    tm = "time %02u:%02u:%02u" % (ttime.tm_hour, ttime.tm_min, ttime.tm_sec)
    print(tm)
    return ttime.tm_hour
def ifrun():
    timehour = int(get_webservertime())
    if timehour in (9, 11):
        print('在第1运行时间内')
    elif timehour in (15, 17):
        print('在第2运行时间内')
    elif timehour in (21, 23):
        print('在第3运行时间内')
    else:
        print('不在运行时间内')
        os._exit(0)


#确认框模板
def chosey(root, list):
    list.append('是')
    root.destroy()
def chosen(root, list):
    list.append('否')
    root.destroy()
def getvalue_yandn(ttext):
    truevalue = []
    root = Tk()
    root.resizable(False, False)
    root.title("")
    mainframe = ttk.Frame(root)
    mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
    screenwidth = root.winfo_screenwidth()
    screenheight = root.winfo_screenheight()
    root.geometry('%sx100+%s+%s' % (len(ttext)*8+100, screenwidth - 850, screenheight - 600))
    Label(mainframe, text='').grid(column=0,row=0)
    Label(mainframe, text='   ').grid(column=0, rowspan=3)
    ttk.Label(mainframe, text=ttext).grid(column=1, row=1, sticky=W)
    ttk.Button(mainframe, text='是', width=5, command=lambda: chosey(root, truevalue)).grid(column=1, row=2, sticky=W)
    ttk.Button(mainframe, text='否', width=5, command=lambda: chosen(root, truevalue)).grid(column=1, row=2, sticky=E)
    root.mainloop()
    return truevalue

#断点续传
    #建立断点续传记录文件
def resume1(pointname):
    pointpath1 = docpath + '\\' + pointname + '\\'
    if os.path.exists(pointpath1) == False:
        os.makedirs(pointpath1)
    pointpath2 = pointpath1 + pointname + 'point.txt'
    f = open(pointpath2, 'w')
    f.close()
    #记录断点
def resume2(url):
    pointpath = docpath + '\\' + infovalue1[0] + '\\' + infovalue1[0] + 'point.txt'
    f = open(pointpath, 'a')
    f.write(url + '\n')
    f.close()
    #读取断点
def resume3():
    pointpath = docpath + '\\' + infovalue1[0] + '\\' + infovalue1[0] + 'point.txt'
    f = open(pointpath, 'r')
    x = f.readlines()
    xx = x[-1].rstrip('\n')
    f.close()
    return xx

#建立爬取目录
def getvalue_build():
    truevalue = []
    root = Tk()
    root.resizable(False, False)
    root.title("初始设置")
    mainframe = ttk.Frame(root)
    mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
    screenwidth = root.winfo_screenwidth()
    screenheight = root.winfo_screenheight()
    root.geometry('300x200+%s+%s' % (screenwidth - 850, screenheight - 600))
    Label(mainframe, text='').grid(column=0, columnspan=4, row=2)
    ttk.Label(mainframe, text='给你的要爬取的文件设置磁盘地址(如D:\或D:\Spider\):').grid(column=0, row=3)
    var = StringVar()
    ttk.Entry(mainframe, width=30, textvariable=var).grid(column=0, row=4)
    ttk.Label(mainframe, text='给你保存文件的文件夹取名(中文/英文):').grid(column=0, row=5)
    var2 = StringVar()
    ttk.Entry(mainframe, width=30, textvariable=var2).grid(column=0, row=6)
    ttk.Label(root, text='').grid(column=0, columnspan=4, row=7)
    ttk.Label(root, text='确认无误后请点击确定').grid(column=0, columnspan=4, row=8)
    ttk.Button(root, text='确定', command=lambda: root.destroy()).grid(column=0, columnspan=4, row=9)
    root.mainloop()
    x1 = var.get()
    x2 = var2.get()
    truevalue.append(x1)
    truevalue.append(x2)
    return truevalue
def builddoc():
    list = ['\\', '/', ':', '*', '?', '"', '<', '>', '|']
    finalpath = []
    enble = True
    while enble == True:
        try:
            if os.path.exists(doctxt) == True:
                try:
                    buildinfo = []
                    openpath = openpyxl.load_workbook(doctxt)
                    sheet = openpath.get_sheet_names()
                    ws = openpath.get_sheet_by_name(sheet[0])
                    info1 = ws.cell(row=1, column=1).value
                    buildinfo.append(info1)
                    if not ':\\' in info1:
                        print('地址读取错误')
                        sys.exit()
                    info2 = ws.cell(row=2, column=1).value
                    buildinfo.append(info2)
                    for each in info2:
                        for un in list:
                            if un in each:
                                print('名字存在非法字符,请重新设置')
                                sys.exit()
                    openpath.save(doctxt)
                except:
                    buildinfo = []
                    print('读取或输入有误,请重新设置项目地址')
                    openpath = openpyxl.load_workbook(doctxt)
                    sheet = openpath.get_sheet_names()
                    ws = openpath.get_sheet_by_name(sheet[0])
                    info = getvalue_build()
                    docpath = info[0]
                    if not ':\\' in docpath:
                        print('地址输入错误')
                        sys.exit()
                    if '\\' not in docpath[-1]:
                        docpath = str(docpath) + '\\'
                    docpath = str(docpath).replace("\\", "\\\\")
                    ws.cell(row=1, column=1, value=docpath)
                    buildinfo.append(docpath)
                    docname = info[1]
                    for each in docname:
                        for un in list:
                            if un in each:
                                print('名字存在非法字符,请重新设置')
                                sys.exit()
                    buildinfo.append(docname)
                    ws.cell(row=2, column=1, value=docname)
                    openpath.save(doctxt)
            else:
                buildinfo = []
                print('新建立项目')
                workbook = xlsxwriter.Workbook(doctxt)
                workbook.add_worksheet()
                workbook.add_worksheet('item')
                workbook.close()
                openpath = openpyxl.load_workbook(doctxt)
                sheet = openpath.get_sheet_names()
                ws = openpath.get_sheet_by_name(sheet[0])
                info = getvalue_build()
                docpath = info[0]
                if not ':\\' in docpath:
                    print('地址输入错误')
                    sys.exit()
                if '\\' not in docpath[-1]:
                    docpath = str(docpath) + '\\'
                docpath = str(docpath).replace("\\", "\\\\")
                ws.cell(row=1, column=1, value=docpath)
                buildinfo.append(docpath)
                docname = info[1]
                for each in docname:
                    for un in list:
                        if un in each:
                            print('名字存在非法字符,请重新设置')
                            sys.exit()
                buildinfo.append(docname)
                ws.cell(row=2, column=1, value=docname)
                openpath.save(doctxt)
            path = str(buildinfo[0]).replace('\n', '')
            name = buildinfo[1]
            finalpath = str(path) + str(name)
            enble = False
        except:
            enble = True
    if os.path.exists(finalpath) == True:
        print('路径已设置')
    else:
        os.makedirs(finalpath)
        print('路径已创建')
    return finalpath

docpath = builddoc()
print(docpath)

#新手帮助
def check(judge):
    judge.append('不再查看')
def getvalue_newer():
    root = Tk()
    judge = []
    root.resizable(False, False)
    root.title("新手帮助")
    mainframe = ttk.Frame(root)
    mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
    screenwidth = root.winfo_screenwidth()
    screenheight = root.winfo_screenheight()
    root.geometry('400x290+%s+%s' % (screenwidth - 850, screenheight - 600))
    ttk.Label(mainframe, text='    ').grid(column=0, rowspan=10)
    ttk.Label(mainframe, text='1.本程序可抓取文字/数据/页面链接/下载图片,如果有多页,也可以翻').grid(column=1, columnspan=3, row=0, sticky=W)
    ttk.Label(mainframe, text='页进行抓取,也可以抓取页面中(如新闻标题)点进去的详情页的信息').grid(column=1, columnspan=3, row=1, sticky=W)
    ttk.Label(mainframe, text='2.页面分为静态页面和动态页面,翻页但网址不变的即为动态页面,只有').grid(column=1, columnspan=3, row=2, sticky=W)
    ttk.Label(mainframe, text='使用自定义流程的方式才能进行抓取').grid(column=1, columnspan=3, row=3, sticky=W)
    ttk.Label(mainframe, text='3.抓取主要是通过粘贴xpath来进行;登录功能以及自定义流程需要安装').grid(column=1, columnspan=3, row=4, sticky=W)
    ttk.Label(mainframe, text='Chrome浏览器,具体见附件<Xpath操作><Chrome驱动设置>').grid(column=1, columnspan=3, row=5, sticky=W)
    ttk.Label(mainframe, text='4.本程序可实现自动化,每个项目结束后可选择是否保存操作行为,之后').grid(column=1, columnspan=3, row=6, sticky=W)
    ttk.Label(mainframe, text='可加载相应醒目的操作,免去重复操作的繁琐').grid(column=1, columnspan=3, row=7,sticky=W)
    ttk.Label(mainframe, text="5.与程序同名的Txt文件中可选择是否自动化,若第三行字段为'自动化'").grid(column=1, columnspan=3, row=8, sticky=W)
    ttk.Label(mainframe, text='则视为开启自动化,将自行加载所有保存过的项目操作').grid(column=1, columnspan=3, row=9,sticky=W)
    ttk.Label(mainframe, text='6.本程序为半可视化,有些响应会出现在黑框中,请按照指示进行操作').grid(column=1, columnspan=3, row=10, sticky=W)
    ttk.Label(mainframe, text='    ').grid(column=0, row=11)
    ttk.Button(mainframe, text='确定', command=lambda: root.destroy()).grid(column=1, columnspan=3, row=12)
    ttk.Checkbutton(mainframe, text='不再查看', command=lambda: check(judge)).grid(column=3, row=12)
    root.mainloop()
    return judge
def newer():
    openpath = openpyxl.load_workbook(doctxt)
    sheet = openpath.get_sheet_names()
    ws = openpath.get_sheet_by_name(sheet[0])
    x = ws.cell(row=4, column=1).value
    if x == '不再查看':
        pass
    else:
        print('请查看新手帮助')
        judge = getvalue_newer()
        if judge == []:
            judge = ['']
        ws.cell(row=4, column=1, value=judge[0])
        openpath.save(doctxt)

#是否加载项目
def getvalue_loadact():
    root = Tk()
    root.resizable(False, False)
    root.title("加载项目")
    mainframe = ttk.Frame(root)
    mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
    screenwidth = root.winfo_screenwidth()
    screenheight = root.winfo_screenheight()
    root.geometry('270x100+%s+%s' % (screenwidth - 850, screenheight - 600))
    ttk.Label(mainframe, text='    ').grid(column=0, row=0)
    ttk.Label(mainframe, text='请输入曾建立的项目名称加载操作信息').grid(column=1, columnspan=2, row=0)
    var = StringVar()
    ttk.Entry(mainframe, width=30, textvariable=var).grid(column=1, columnspan=2, row=1)
    ttk.Label(mainframe, text='    ').grid(column=0, row=2)
    ttk.Button(mainframe, text='确定', command=lambda: root.destroy()).grid(column=1, columnspan=2, row=3)
    root.mainloop()
    x = var.get()
    return x

def loadaction():
    enble = True
    getvalue = []
    judgeload = ''
    list = ['\\', '/', ':', '*', '?', '"', '<', '>', '|']
    while enble == True:
        try:
            tt = '是否要加载保存过操作记录的项目?'
            xx = getvalue_yandn(tt)
            ifload = xx[0]
            if ifload == '是':
                pastname = getvalue_loadact()
                for each in pastname:
                    for un in list:
                        if un in each:
                            print('名字存在非法字符,请重新设置')
                            sys.exit()
                actpath = docpath + '\\' + pastname + '\\' + pastname + '.xlsx'
                if os.path.exists(actpath) == False:
                    print('该项目不存在或未保存操作记录,请重新输入')
                    sys.exit()
                else:
                    openpath = openpyxl.load_workbook(actpath)
                    sheet = openpath.get_sheet_names()
                    ws = openpath.get_sheet_by_name(sheet[0])
                    cx = ws.max_column
                    value1 = []
                    for i in range(0, 16):
                        x = ws.cell(row=i + 1, column=1).value
                        value1.append(x)
                    getvalue.append(value1)
                    for i in range(2, int(cx)+1):
                        value2 = []
                        x1 = ws.cell(row = 1, column=i).value
                        x2 = ws.cell(row=2, column=i).value
                        x3 = ws.cell(row=3, column=i).value
                        value2.append(x1)
                        value2.append(x2)
                        value2.append(x3)
                        getvalue.append(value2)
                judgeload = ifload + judgeload
                infovalue1[0] = pastname
                enble = False
            else:
                enble = False
        except:
            enble = True
    lens = len(getvalue)
    for i in range(0, lens):
        infovalue3.append([])
        infovalue3[i] = getvalue[i]
    return judgeload

#是否保存项目操作
def getvalue_saveact():
    root = Tk()
    root.resizable(False, False)
    root.title("建立项目")
    mainframe = ttk.Frame(root)
    mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
    screenwidth = root.winfo_screenwidth()
    screenheight = root.winfo_screenheight()
    root.geometry('300x100+%s+%s' % (screenwidth - 850, screenheight - 600))
    ttk.Label(mainframe, text='    ').grid(column=0, row=0)
    ttk.Label(mainframe, text='请为该次项目取名,将记录操作信息以实现自动化:').grid(column=1, columnspan=2, row=0)
    var = StringVar()
    ttk.Entry(mainframe, width=30, textvariable=var).grid(column=1, columnspan=2, row=1)
    ttk.Label(mainframe, text='    ').grid(column=0, row=2)
    ttk.Button(root, text='确定', command=lambda: root.destroy()).grid(column=0, row=3)
    root.mainloop()
    x = var.get()
    return x
def buildaction(actioninfo):
    list = ['\\', '/', ':', '*', '?', '"', '<', '>', '|']
    len1 =len(infovalue2)
    enble = True
    while enble == True:
        try:
            if actioninfo == '':
                enble = False
            else:
                for each in actioninfo:
                    for un in list:
                        if un in each:
                            print('名字存在非法字符,请重新设置')
                            sys.exit()
                actpath = docpath + '\\' + actioninfo
                if os.path.exists(actpath) == True:
                    tt = '该路径已存在,是否更新操作记录?(是则更新/否则重新输入)'
                    xx = getvalue_yandn(tt)
                    x = xx[0]
                    if x == '是':
                        path = actpath + '\\' + actioninfo + '.xlsx'
                        workbook = xlsxwriter.Workbook(path)
                        workbook.close()
                        openpath = openpyxl.load_workbook(path)
                        sheet = openpath.get_sheet_names()
                        ws = openpath.get_sheet_by_name(sheet[0])
                        for i in range(0, len1):
                            len2 = len(infovalue2[i])
                            for ii in range(0, len2):
                                ws.cell(row=ii + 1, column=i + 1, value=infovalue2[i][ii])
                        openpath.save(path)
                        infovalue1[0] = actioninfo
                        enble = False
                    else:
                        actioninfo2 = getvalue_saveact()
                        actpath2 = docpath + '\\' + actioninfo2
                        while os.path.exists(actpath2) == True:
                            print('项目已存在,请重新输入')
                            actioninfo2 = getvalue_saveact()
                            actpath2 = docpath + '\\' + actioninfo2
                        os.makedirs(actpath2)
                        path = actpath2 + '\\' + actioninfo2 + '.xlsx'
                        workbook = xlsxwriter.Workbook(path)
                        workbook.close()
                        openpath = openpyxl.load_workbook(path)
                        sheet = openpath.get_sheet_names()
                        ws = openpath.get_sheet_by_name(sheet[0])
                        for i in range(0, len1):
                            len2 = len(infovalue2[i])
                            for ii in range(0, len2):
                                ws.cell(row=ii + 1, column=i + 1, value=infovalue2[i][ii])
                        openpath.save(path)
                        infovalue1[0] = actioninfo2
                        enble = False
                else:
                    os.makedirs(actpath)
                    path = actpath + '\\' + actioninfo + '.xlsx'
                    workbook = xlsxwriter.Workbook(path)
                    workbook.close()
                    openpath = openpyxl.load_workbook(path)
                    sheet = openpath.get_sheet_names()
                    ws = openpath.get_sheet_by_name(sheet[0])
                    for i in range(0, len1):
                        len2 = len(infovalue2[i])
                        for ii in range(0, len2):
                            ws.cell(row=ii + 1, column=i + 1, value=infovalue2[i][ii])
                    openpath.save(path)
                    infovalue1[0] = actioninfo
                    enble = False
        except:
            enble = True

#登录网页
    #获取验证码
def getidencode(location, size, imagename):
    from PIL import Image
    left = location['x']
    top = location['y']
    right = location['x'] + size['width']
    bottom = location['y'] + size['height']
    im = Image.open(imagename)
    im = im.convert('L')
    box = (left, top, right, bottom)
    im = im.crop(box)
    im.save(imagename)
    time.sleep(3)
    img = Image.open(imagename)
    time.sleep(3)
    idencode = pytesseract.image_to_string(img)
    return idencode
    #获取cookie
def getvalue_login(setinfo):
    truevalue = []
    root = Tk()
    root.resizable(False, False)
    root.title("登录信息")
    screenwidth = root.winfo_screenwidth()
    screenheight = root.winfo_screenheight()
    root.geometry('300x430+%s+%s' % (screenwidth - 850, screenheight - 600))
    l2 = Label(root, text='*用户名输入框路径:')
    l2.pack()
    var2 = StringVar()
    var2.set(setinfo[1])
    Entry(root, width=40, textvariable=var2).pack()
    l3 = Label(root, text='*密码输入框路径:')
    l3.pack()
    var3 = StringVar()
    var3.set(setinfo[2])
    Entry(root, width=40, textvariable=var3).pack()
    l4 = Label(root, text='验证码输入框路径(无验证码则不填):')
    l4.pack()
    var4 = StringVar()
    var4.set(setinfo[3])
    Entry(root, width=40, textvariable=var4).pack()
    l5 = Label(root, text='验证码图片路径(无验证码则不填/可不填):')
    l5.pack()
    var5 = StringVar()
    var5.set(setinfo[4])
    Entry(root, width=40, textvariable=var5).pack()
    l6 = Label(root, text='*登录按键路径:')
    l6.pack()
    var6 = StringVar()
    var6.set(setinfo[5])
    Entry(root, width=40, textvariable=var6).pack()
    l7 = Label(root, text='*用户名(无需在打开的网页中输入):')
    l7.pack()
    var7 = StringVar()
    var7.set(setinfo[6])
    Entry(root, width=20, textvariable=var7).pack()
    l8 = Label(root, text='*密码(无需在打开的网页中输入):')
    l8.pack()
    var8 = StringVar()
    var8.set(setinfo[7])
    Entry(root, width=20, textvariable=var8).pack()
    l9 = Label(root, text='验证码输入(无验证码则不填/验证码难以识别时可填):')
    l9.pack()
    var9 = StringVar()
    var9.set(setinfo[8])
    Entry(root, width=20, textvariable=var9).pack()
    Button(root, text='确定', width=10, command=lambda :root.destroy()).pack()
    Label(root, text='确认输入无误后点击确定').pack()
    root.mainloop()
    truevalue2 = var2.get()
    truevalue3 = var3.get()
    truevalue4 = var4.get()
    truevalue5 = var5.get()
    truevalue6 = var6.get()
    truevalue7 = var7.get()
    truevalue8 = var8.get()
    truevalue9 = var9.get()
    truevalue.append(truevalue2)
    truevalue.append(truevalue3)
    truevalue.append(truevalue4)
    truevalue.append(truevalue5)
    truevalue.append(truevalue6)
    truevalue.append(truevalue7)
    truevalue.append(truevalue8)
    truevalue.append(truevalue9)
    return truevalue
def getvalue_afterurl():
    root = Tk()
    root.resizable(False, False)
    root.title("登录后的网页")
    mainframe = ttk.Frame(root)
    mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
    screenwidth = root.winfo_screenwidth()
    screenheight = root.winfo_screenheight()
    root.geometry('270x130+%s+%s' % (screenwidth - 850, screenheight - 600))
    ttk.Label(mainframe, text='    ').grid(column=0, row=0)
    ttk.Label(mainframe, text='请输入登录后的网页地址\n(成功登录后输入,失败请直接按确定继续)').grid(column=1, columnspan=2, row=0)
    var = StringVar()
    ttk.Entry(mainframe, width=30, textvariable=var).grid(column=1, columnspan=2, row=1)
    ttk.Label(mainframe, text='    ').grid(column=0, row=2)
    ttk.Button(mainframe, text='确定', command=lambda: root.destroy()).grid(column=1, columnspan=2, row=3)
    root.mainloop()
    x = var.get()
    return x
def getvalue_logurl():
    root = Tk()
    root.resizable(False, False)
    root.title("登录网址")
    mainframe = ttk.Frame(root)
    mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
    screenwidth = root.winfo_screenwidth()
    screenheight = root.winfo_screenheight()
    root.geometry('270x110+%s+%s' % (screenwidth - 850, screenheight - 600))
    ttk.Label(mainframe, text='    ').grid(column=0, row=0)
    ttk.Label(mainframe, text='请输入登录界面的网页地址').grid(column=1, columnspan=2, row=0)
    var = StringVar()
    ttk.Entry(mainframe, width=30, textvariable=var).grid(column=1, columnspan=2, row=1)
    ttk.Label(mainframe, text='    ').grid(column=0, row=2)
    ttk.Button(mainframe, text='确定', command=lambda: root.destroy()).grid(column=1, columnspan=2, row=3)
    root.mainloop()
    x = var.get()
    return x
def login(judgeload):
    loginfo = []
    if judgeload == '是':
        pastname = infovalue1[0]
        putpath = docpath + '\\' + pastname + '\\' + pastname + '.xlsx'
        openpath = openpyxl.load_workbook(putpath)
        sheet = openpath.get_sheet_names()
        ws = openpath.get_sheet_by_name(sheet[0])
        info1 = ws.cell(row=2, column=1).value
        loginfo.append(info1)
        infovalue2[0][1] = info1
        info2 = ws.cell(row=3, column=1).value
        loginfo.append(info2)
        infovalue2[0][2] = info2
        info3 = ws.cell(row=4, column=1).value
        loginfo.append(info3)
        infovalue2[0][3] = info3
        info4 = ws.cell(row=5, column=1).value
        loginfo.append(info4)
        infovalue2[0][4] = info4
        info5 = ws.cell(row=6, column=1).value
        loginfo.append(info5)
        infovalue2[0][5] = info5
        info6 = ws.cell(row=7, column=1).value
        loginfo.append(info6)
        infovalue2[0][6] = info6
        info7 = ws.cell(row=8, column=1).value
        loginfo.append(info7)
        infovalue2[0][7] = info7
        info8 = ws.cell(row=9, column=1).value
        loginfo.append(info8)
        infovalue2[0][8] = info8
        info9 = ''
        loginfo.append(info9)
        info10 = ws.cell(row=10, column=1).value
        infovalue2[0][9] = info10
        drive = webdriver.Chrome()
        drive.get(info1)
        print('自动登录时验证码为自动识别,若识别率差可以在网页上手动输入;有其他错误可手动刷新网页')
    else:
        setinfo1 = ['','','','','','','','','']
        info1 = getvalue_logurl()
        loginfo.append(info1)
        if 'http' not in info1:
            info1 = 'http://' + info1
        info1 = str(info1).replace(' ', '')
        drive = webdriver.Chrome()
        drive.get(info1)
        info = getvalue_login(setinfo1)
        loginfo += info
        info2 = loginfo[1]
        info3 = loginfo[2]
        info4 = loginfo[3]
        info5 = loginfo[4]
        info6 = loginfo[5]
        info7 = loginfo[6]
        info8 = loginfo[7]
        info9 = loginfo[8]
    enble = True
    usernamepath = info2
    passwordpath = info3
    idencodepath = info4
    idencodeimgpath = info5
    subpath = info6
    usernameinput = info7
    passwordinput = info8
    while enble == True:
        try:
            usernamename = drive.find_element_by_xpath(str(usernamepath)).get_attribute('name')
            passwordname = drive.find_element_by_xpath(str(passwordpath)).get_attribute('name')
            drive.find_element_by_name(usernamename).send_keys(str(usernameinput))
            time.sleep(1)
            drive.find_element_by_name(passwordname).send_keys(str(passwordinput))
            if idencodepath != '' and idencodepath != None:
                idencodeinput = info9
                idencodename = drive.find_element_by_xpath(str(idencodepath)).get_attribute('name')
                if idencodeinput != '' and idencodeinput != None:
                    drive.find_element_by_name(idencodename).send_keys(str(idencodeinput))
                else:
                    from PIL import Image
                    try:
                        imgname = 'code.jpg'
                        drive.get_screenshot_as_file(imgname)
                        location = drive.find_element_by_xpath(str(idencodeimgpath)).location
                        size = drive.find_element_by_xpath(str(idencodeimgpath)).size
                        idencode = getidencode(location, size, imgname)
                        drive.find_element_by_name(idencodename).send_keys(idencode)
                    except:
                        continue
            time.sleep(0.5)
            drive.find_element_by_xpath(str(subpath)).click()
            time.sleep(0.5)
            if not judgeload == '是':
                afterinfo = getvalue_afterurl()
                afterinfo = str(afterinfo).replace(' ', '')
                if 'http' not in afterinfo:
                    afterinfo = 'http://' + afterinfo
                infovalue2[0][9] = afterinfo
            afterurl = infovalue2[0][9]
            x = str(drive.current_url)
            if x == afterurl:
                enble = False
            else:
                sys.exit()
        except:
            try:
                drive.switch_to.alert().accept()
            except:
                print('重新登录中...')
                drive.refresh()
            if not judgeload == '是':
                info = getvalue_login(loginfo)
                info2 = info[0]
                info3 = info[1]
                info4 = info[2]
                info5 = info[3]
                info6 = info[4]
                info7 = info[5]
                info8 = info[6]
                info9 = info[7]
                usernamepath = info2
                passwordpath = info3
                idencodepath = info4
                idencodeimgpath = info5
                subpath = info6
                usernameinput = info7
                passwordinput = info8
            time.sleep(1)
            enble = True
    print('登录成功')
    if not judgeload == '是':
        tt = '要保存登录过的信息以便下次登录吗?'
        xx = getvalue_yandn(tt)
        answer = xx[0]
        if answer == '是':
            history = getvalue_saveact()
            newpath = docpath + '\\' + history + '\\' + history + '.xlsx'
            while os.path.exists(newpath) == True:
                print('文件已存在,请重新命名')
                history = getvalue_saveact()
                newpath = docpath + '\\' + history + '\\' + history + '.xlsx'
            else:
                actpath = docpath + '\\' + history
                if os.path.exists(actpath) == False:
                    os.makedirs(actpath)
                workbook = xlsxwriter.Workbook(newpath)
                workbook.close()
                openpath = openpyxl.load_workbook(newpath)
                sheet = openpath.get_sheet_names()
                ws = openpath.get_sheet_by_name(sheet[0])
                ws.cell(row=2, column=1, value=loginfo[0])
                ws.cell(row=3, column=1, value=loginfo[1])
                ws.cell(row=4, column=1, value=loginfo[2])
                ws.cell(row=5, column=1, value=loginfo[3])
                ws.cell(row=6, column=1, value=loginfo[4])
                ws.cell(row=7, column=1, value=loginfo[5])
                ws.cell(row=8, column=1, value=loginfo[6])
                ws.cell(row=9, column=1, value=loginfo[7])
                ws.cell(row=10, column=1, value=infovalue2[0][9])
                infovalue1[0] = history
                openpath.save(newpath)
                print('信息保存完毕')
    else:
        if judgeload == '是':
            tt = '要更新登录的信息吗?'
            xx = getvalue_yandn(tt)
            answer = xx[0]
            if answer == '是':
                pastname = infovalue1[0]
                putpath = docpath + '\\' + pastname + '\\' + pastname + '.xlsx'
                openpath = openpyxl.load_workbook(putpath)
                sheet = openpath.get_sheet_names()
                ws = openpath.get_sheet_by_name(sheet[0])
                ws.cell(row=2, column=1, value=loginfo[0])
                ws.cell(row=3, column=1, value=loginfo[1])
                ws.cell(row=4, column=1, value=loginfo[2])
                ws.cell(row=5, column=1, value=loginfo[3])
                ws.cell(row=6, column=1, value=loginfo[4])
                ws.cell(row=7, column=1, value=loginfo[5])
                ws.cell(row=8, column=1, value=loginfo[6])
                ws.cell(row=9, column=1, value=loginfo[7])
                ws.cell(row=10, column=1, value=infovalue2[0][9])
                print('信息更新完毕')
    for i in range(0, 8):
        infovalue2[0][i+1] = loginfo[i]
    infovalue2[0][1] = loginfo[0]
    fullcookies = drive.get_cookies()
    return fullcookies
#保存cookie
def savecookie(judgeload):
    cookies = login(judgeload)
    pastname = infovalue1[0]
    num = len(cookies)
    cookielist = []
    for i in range(0, int(num)):
        fac = cookies[i]['name'] + '=' + cookies[i]['value'] + ';'
        cookielist.append(fac)
    cookie = ''.join(cookielist)
    cookiepath = docpath + '\\' + pastname + '\\' + pastname + 'cookie.txt'
    f = open(cookiepath, 'w')
    f.writelines(cookie)
    f.close()
    return cookie
#将cookie添加进header
def postheader(judgeload):
    pastname = infovalue1[0]
    cookiepath = docpath + '\\' + pastname + '\\' + pastname + 'cookie.txt'
    postcookie = []
    if os.path.exists(cookiepath) == True:
        cookie = open(cookiepath).readlines()
        postcookie.append(cookie[0])
        if judgeload == '是':
            actpath = docpath + '\\' + pastname + '\\' + pastname + '.xlsx'
            openpath = openpyxl.load_workbook(actpath)
            sheet = openpath.get_sheet_names()
            ws = openpath.get_sheet_by_name(sheet[0])
            for i in range(0, 16):
                infovalue2[0][i] = ws.cell(row=i+1, column=1).value
    else:
        cookie = savecookie(judgeload)
        postcookie.append(cookie)
    header['Cookie'] = postcookie[0]
    return header

#打开网页获取资源
def getsource(url, header):
    req = urllib.request.Request(url, headers=header)
    data = urllib.request.urlopen(req).read()
    chardetx = chardet.detect(data)
    html = requests.get(url, headers=header, allow_redirects=False)
    html.encoding = chardetx['encoding']
    return html

##抓取单页/多页的文字/数据/链接信息
#获取网页内容
    #选择获取的信息类型
def help2():
    root = Tk()
    root.resizable(False, False)
    root.title("注意事项")
    mainframe = ttk.Frame(root)
    mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
    screenwidth = root.winfo_screenwidth()
    screenheight = root.winfo_screenheight()
    root.geometry('400x180+%s+%s' % (screenwidth - 650, screenheight - 500))
    ttk.Label(root, text='1.同一类型的信息每次点击添加键将覆盖,每点击一次,上一次的输入都会保').grid(row=0, sticky=W)
    ttk.Label(root, text='存,输入有误重新点击添加即可,不影响抓取;').grid(row=1, sticky=W)
    ttk.Label(root, text='2.A/B/D的第二栏请按排列的顺序粘贴信息1的下一个信息2的路径,想抓取').grid(row=2, sticky=W)
    ttk.Label(root, text='一整列就粘贴一列中的第2个,想抓取一整行同理;').grid(row=3, sticky=W)
    ttk.Label(root, text='3.C无法抓取一段话首尾的文字,且要截取的信息限50个字符内,输入要截取').grid(row=4, sticky=W)
    ttk.Label(root, text='信息的前后文以便定位,前后文越具体则定位越准确,此处会抓取所有符合定').grid(row=5, sticky=W)
    ttk.Label(root, text='位的信息,抓取后请根据需要筛选.').grid(row=6, sticky=W)
    ttk.Button(root, text='确认', command=lambda: root.destroy()).grid(row=7)
    root.mainloop()
def addvar1(mainroot, truelist, row):
    list = []
    ttk.Label(mainroot, text='信息1内容路径:').grid(column=1, row=row)
    var = StringVar()
    ttk.Entry(mainroot, width=20, textvariable=var).grid(column=2, row=row, sticky=W)
    ttk.Label(mainroot, text='').grid(column=4, row=row)
    ttk.Label(mainroot, text='信息2内容路径(若不批量抓取则不填):').grid(column=5, row=row)
    var2 = StringVar()
    ttk.Entry(mainroot, width=20, textvariable=var2).grid(column=6, row=row, sticky=W)
    mainroot.mainloop()
    x1 = var.get()
    x2 = var2.get()
    list.append('X')
    list.append(x1)
    list.append(x2)
    truelist.append(list)
def addvar2(mainroot, truelist, row):
    list = []
    ttk.Label(mainroot, text='截取信息的前文:').grid(column=1, row=row+1)
    var = StringVar()
    ttk.Entry(mainroot, width=20, textvariable=var).grid(column=2, row=row+1, sticky=W)
    ttk.Label(mainroot, text='').grid(column=4, row=row+1)
    ttk.Label(mainroot, text='截取信息的后文:').grid(column=5, row=row+1)
    var2 = StringVar()
    ttk.Entry(mainroot, width=20, textvariable=var2).grid(column=6, row=row+1, sticky=W)
    mainroot.mainloop()
    x1 = var.get()
    x2 = var2.get()
    list.append('Z')
    list.append(x1)
    list.append(x2)
    truelist.append(list)
def addvar3(mainroot, truelist, row):
    list = []
    ttk.Label(mainroot, text='链接1内容路径:').grid(column=1, row=row+2)
    var = StringVar()
    ttk.Entry(mainroot, width=20, textvariable=var).grid(column=2, row=row+2, sticky=W)
    ttk.Label(mainroot, text='').grid(column=4, row=row+2)
    ttk.Label(mainroot, text='链接2内容路径(若不批量抓取则不填):').grid(column=5, row=row+2)
    var2 = StringVar()
    ttk.Entry(mainroot, width=20, textvariable=var2).grid(column=6, row=row+2, sticky=W)
    mainroot.mainloop()
    x1 = var.get()
    x2 = var2.get()
    list.append('U')
    list.append(x1)
    list.append(x2)
    truelist.append(list)
def addvar4(mainroot, truelist, row):
    list = []
    ttk.Label(mainroot, text='图片1下载路径:').grid(column=1, row=row+3)
    var = StringVar()
    ttk.Entry(mainroot, width=20, textvariable=var).grid(column=2, row=row+3, sticky=W)
    ttk.Label(mainroot, text='').grid(column=4, row=row+3)
    ttk.Label(mainroot, text='图片2下载路径(若不批量抓取则不填):').grid(column=5, row=row+3)
    var2 = StringVar()
    ttk.Entry(mainroot, width=20, textvariable=var2).grid(column=6, row=row+3, sticky=W)
    mainroot.mainloop()
    x1 = var.get()
    x2 = var2.get()
    list.append('M')
    list.append(x1)
    list.append(x2)
    truelist.append(list)
def getvalue_spider():
    truevalue = []
    row = 3
    root = Tk()
    root.resizable(False, False)
    root.title("抓取设置")
    screenwidth = root.winfo_screenwidth()
    screenheight = root.winfo_screenheight()
    mainframe = ttk.Frame(root)
    root.geometry('750x270+%s+%s' % (screenwidth - 850, screenheight - 600))
    mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
    ttk.Label(mainframe,text="以新闻'NBA总决赛现场报道'为例,若要抓取整个标题文字选择A;若要截取其中的文字例如'总决赛'则选择B;若要抓取新闻的详情页链接则选择C;\n"
                             "若要下载该新闻的图则选择D.每次点击视为新的添加,可不限次数添加.").grid(columnspan=7)
    ttk.Button(mainframe, width=15, text='注意事项', command=lambda: help2()).grid(column=0, row=2, sticky=W)
    ttk.Button(mainframe, width=15, text="A.添加抓取方式1", command=lambda: addvar1(mainframe, truevalue, row)).grid(column=0, row=row, sticky=W)
    ttk.Button(mainframe, width=15, text="B.添加抓取方式2", command=lambda: addvar2(mainframe, truevalue, row)).grid(column=0, row=row+1, sticky=W)
    ttk.Button(mainframe, width=15, text="C.添加抓取链接", command=lambda: addvar3(mainframe, truevalue, row)).grid(column=0, row=row+2, sticky=W)
    ttk.Button(mainframe, width=15, text="D.添加抓取图片", command=lambda: addvar4(mainframe, truevalue, row)).grid(column=0, row=row+3, sticky=W)
    var = StringVar()
    var2 = StringVar()
    ttk.Entry(mainframe, width=20, textvariable=var).grid(column=2, row=row+4, sticky=W)
    ttk.Entry(mainframe, width=10, textvariable=var2).grid(column=6, row=row+4, sticky=W)
    ttk.Label(mainframe, text='若需翻页请粘贴<下一页>按键的路径:').grid(columnspan=2, row=row+4, sticky=W)
    ttk.Label(mainframe, text='').grid(column=4, row=row+4)
    ttk.Label(mainframe, text='需要抓取几页:').grid(column=5, row=row+4)
    ttk.Label(mainframe, text='').grid(row=row+5)
    ttk.Label(root, text='添加完毕后请点击确认').grid(row=row+6)
    ttk.Button(root, text='确认', command=lambda: root.destroy()).grid(row=row+7)
    root.mainloop()
    x1 = var.get()
    x2 = var2.get()
    x = []
    x.append(x1)
    x.append(x2)
    truevalue.append(x)
    return truevalue
def whichway():
    ways = getvalue_spider()
    return ways
    #一个resultinfo是一类信息
def choseway(ways, url, header, auto):
    allresultinfo = []
    if not auto[0] == '是':
        infovalue2[0][10] = ways[-1][0]
        infovalue2[0][11] = ways[-1][1]
        del ways[-1]
    for each in ways:
        resultinfo = []
        html = getsource(url, header)
        infovalue2.append(each)
        try:
            if each[0] == 'Z':
                html = html.text
                content = BeautifulSoup(html, 'lxml')
                resultinfo = judgezcontent(each, content)
            elif each[0] == 'U':
                html = html.content
                content = etree.HTML(html)
                resultinfo = judgexurl(each, content)
            elif each[0] == 'M':
                html = html.content
                content = etree.HTML(html)
                resultinfo = judgeimg(each, content)
            elif each[0] == 'X':
                html = html.content
                content = etree.HTML(html)
                resultinfo = judgexcontent(each, content)
            else:
                pass
            allresultinfo.append(resultinfo)
        except:
            continue
    return allresultinfo
def pathagain():
    root = Tk()
    root.resizable(False, False)
    root.title("路径")
    mainframe = ttk.Frame(root)
    mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
    screenwidth = root.winfo_screenwidth()
    screenheight = root.winfo_screenheight()
    root.geometry('300x100+%s+%s' % (screenwidth - 850, screenheight - 600))
    ttk.Label(mainframe, text='    ').grid(column=0, row=0)
    ttk.Label(mainframe, text='两个信息路径一样,无法获取多个信息,请重新输入:').grid(column=1, columnspan=2, row=0)
    var = StringVar()
    ttk.Entry(mainframe, width=30, textvariable=var).grid(column=1, columnspan=2, row=1)
    ttk.Label(mainframe, text='    ').grid(column=0, row=2)
    ttk.Button(root, text='确定', command=lambda: root.destroy()).grid(column=0, row=3)
    root.mainloop()
    x = var.get()
    return x
    #xpath信息判断
def judgexcontent(each, content):
    result = ''
    resultinfo = ['X']
    trace = each[1].replace('/text()', '')
    text = content.xpath(str(trace) + '/text()')
    if text == []:
        trace = trace.replace('/tbody', '')
        text = content.xpath(str(trace) + '/text()')
    lentext = len(text)
    if lentext == 0:
        print('无法获取信息')
        sys.exit()
    else:
        judge = ''
        if each[2] != '' and each[2] != None:
            judge = '是'
        if judge == '是':
            trace2 = each[2]
            text = content.xpath(str(trace2) + '/text()')
            if text == []:
                trace2 = trace2.replace('/tbody', '')
            if trace2 == trace:
                if not lentext == 1:
                    print('该内容无法精确提取,抓取后请根据需要筛选')
                    result = result + 'result1'
                    resultinfo.append(result)
                    resultinfo.append(trace)
                else:
                    while trace2 == trace:
                        print('两个信息路径一样,无法获取多个信息,请重新输入')
                        trace2 = pathagain()
                        each[2] = trace2
                        text = content.xpath(str(trace2) + '/text()')
                        if text == []:
                            trace2 = trace2.replace('/tbody', '')
                    if trace2 == '' or trace2 == None:
                        result = result + 'result3'
                        resultinfo.append(result)
                        resultinfo.append(trace)
                    else:
                        result = result + 'result2'
                        resultinfo.append(result)
                        resultinfo.append(trace)
                        resultinfo.append(trace2)
            else:
                if lentext == 1:
                     result = result + 'result2'
                     resultinfo.append(result)
                     resultinfo.append(trace)
                     resultinfo.append(trace2)
                else:
                    print('该内容可能无法完整提取,抓取后请根据需要筛选')
                    result = result + 'result4'
                    resultinfo.append(result)
                    resultinfo.append(trace)
                    resultinfo.append(trace2)
        else:
            result = result + 'result3'
            resultinfo.append(result)
            resultinfo.append(trace)
    return resultinfo
    #利用xpath获取数据或文字内容
def getxcontent(content, resultinfo):
    info = []
    trace = resultinfo[2]
    text = content.xpath(str(trace) + '/text()')
    lentext = len(text)
    if resultinfo[1] == 'result1':
        for i in range(0, lentext):
            info.append(text[i])
        print('信息抓取中...')
    elif resultinfo[1] == 'result2':
        trace2 = resultinfo[3]
        enble = True
        num = []
        lens = len(trace)
        try:
            for x1 in range(0, lens):
                if trace[x1] not in trace2[x1]:
                    if trace[x1].isdigit() == True:
                        num.append(trace[x1])
            for x2 in range(0, lens):
                if trace2[x2] not in trace[x2]:
                    if trace2[x2].isdigit() == True:
                        num.append(trace2[x2])
            margin = int(num[1]) - int(num[0])
            n = int(num[0])
            while enble == True:
                lasttrace = ''
                for x in range(0, lens):
                    if trace[x] in trace2[x]:
                        lasttrace = lasttrace + str(trace[x])
                    else:
                        lasttrace = lasttrace + str(n)
                if content.xpath(str(lasttrace).rstrip('/')):
                    text = content.xpath(str(lasttrace).rstrip('/') + '/text()')
                    if text == []:
                        text = ''
                    else:
                        text = text[0].strip()
                    info.append(text)
                    n += margin
                    enble = True
                else:
                    print('信息抓取中...')
                    enble = False
        except:
            print('抓取信息出错了')
            sys.exit()
    elif resultinfo[1] == 'result3':
        text = text[0].strip()
        info.append(text)
        print('信息抓取中...')
    elif resultinfo[1] == 'result4':
        trace2 = resultinfo[3]
        enble = True
        num = []
        lens = len(trace)
        try:
            for x1 in range(0, lens):
                if trace[x1] not in trace2[x1]:
                    if trace[x1].isdigit() == True:
                        num.append(trace[x1])
            for x2 in range(0, lens):
                if trace2[x2] not in trace[x2]:
                    if trace2[x2].isdigit() == True:
                        num.append(trace2[x2])
            margin = int(num[1]) - int(num[0])
            n = int(num[0])
            while enble == True:
                lasttrace = ''
                for x in range(0, lens):
                    if trace[x] in trace2[x]:
                        lasttrace = lasttrace + str(trace[x])
                    else:
                        lasttrace = lasttrace + str(n)
                if content.xpath(str(lasttrace).rstrip('/')):
                    text = content.xpath(str(lasttrace).rstrip('/') + '/text()')
                    if text == []:
                        text1 = ''
                    else:
                        lent = len(text)
                        text1 = ''
                        for i in range(0, lent):
                            text1 = text1 + text[i].strip()
                    info.append(text1)
                    n += margin
                    enble = True
                else:
                    print('信息抓取中...')
                    enble = False
        except:
            print('抓取信息出错了')
            sys.exit()
    else:
        print('抓取信息出错了')
    return info
    #xpath链接判断
def judgexurl(each, content):
    result = ''
    resultinfo = ['U']
    trace = each[1]
    if not re.search('/a', str(trace)):
        print('该内容不包含链接')
        sys.exit()
    else:
        if re.search('/a.*?/', str(trace)):
            mould = re.search('/a.*?/', str(trace)).group()
            place = trace.index(mould)
            trace = trace.replace(str(trace[place:]), str(mould))
        elif re.search('/a/', str(trace)):
            place = trace.index('/a/')
            trace = trace.replace(str(trace[place:]), '/a')
    text = content.xpath(str(trace).rstrip('/') + '/@href')
    if text == []:
        trace = trace.replace('/tbody', '')
        text = content.xpath(str(trace) + '/@href')
    lentext = len(text)
    if not lentext == 1:
        if lentext == 0:
            print('无法获取信息')
            sys.exit()
        else:
            print('该内容无法精确提取,抓取后请根据需要筛选')
            result = result + 'result1'
            resultinfo.append(result)
            resultinfo.append(trace)
    else:
        text = content.xpath(str(trace) + '/@href')[0].strip()
        if re.search('javascript:', text):
            print('该链接无法抓取')
            sys.exit()
        judge = ''
        if each[2] != '' and each[2] != None:
            judge = '是'
        if judge == '是':
            trace2 = each[2]
            if not re.search('/a', str(trace2)):
                print('该内容不包含链接')
                sys.exit()
            else:
                if re.search('/a.*?/', str(trace2)):
                    mould = re.search('/a.*?/', str(trace2)).group()
                    place = trace2.index(mould)
                    trace2 = trace2.replace(str(trace2[place:]), str(mould))
                elif re.search('/a/', str(trace2)):
                    place = trace2.index('/a/')
                    trace2 = trace2.replace(str(trace2[place:]), '/a')
            text = content.xpath(str(trace2).rstrip('/') + '/@href')
            if text == []:
                trace2 = trace2.replace('/tbody', '')
            while trace2 == trace:
                print('两个链接路径一样,无法获取多个链接,请重新输入')
                trace2 = pathagain()
                each[2] = trace2
                if not re.search('/a', str(trace2)):
                    print('该内容不包含链接')
                    sys.exit()
                else:
                    if re.search('/a.*?/', str(trace2)):
                        mould = re.search('/a.*?/', str(trace2)).group()
                        place = trace2.index(mould)
                        trace2 = trace2.replace(str(trace2[place:]), str(mould))
                    elif re.search('/a/', str(trace2)):
                        place = trace2.index('/a/')
                        trace2 = trace2.replace(str(trace2[place:]), '/a')
                text = content.xpath(str(trace2).rstrip('/') + '/@href')
                if text == []:
                    trace2 = trace2.replace('/tbody', '')
            if trace2 == '' or trace2 == None:
                result = result + 'result3'
                resultinfo.append(result)
                resultinfo.append(trace)
            else:
                result = result + 'result2'
                resultinfo.append(result)
                resultinfo.append(trace)
                resultinfo.append(trace2)
        else:
            text = text[0].strip()
            if re.search('javascript:', text):
                print('该链接无法抓取')
                sys.exit()
            result = result + 'result3'
            resultinfo.append(result)
            resultinfo.append(trace)
    return resultinfo
    #利用xpath获取链接
def getxurl(url, content, resultinfo):
    info = []
    protocol, s1 = urllib.parse.splittype(url)
    host, s2 = urllib.parse.splithost(s1)
    trace = resultinfo[2]
    text = content.xpath(str(trace) + '/@href')
    lentext = len(text)
    if resultinfo[1] == 'result1':
        for i in range(0, lentext):
            text = content.xpath(str(trace) + '/@href')
            if text[i][0] == '/':
                text = str(host) + str(text[i])
            elif text[i][0] == 'h':
                text = str(text[i])
            else:
                text = str(host) + '/' + str(text[i])
            info.append(text)
        print('链接抓取中...')
    elif resultinfo[1] == 'result2':
        trace2 = resultinfo[3]
        enble = True
        num = []
        lens = len(trace)
        try:
            for x1 in range(0, lens):
                if trace[x1] not in trace2[x1]:
                    if trace[x1].isdigit() == True:
                        num.append(trace[x1])
            for x2 in range(0, lens):
                if trace2[x2] not in trace[x2]:
                    if trace2[x2].isdigit() == True:
                        num.append(trace2[x2])
            margin = int(num[1]) - int(num[0])
            n = int(num[0])
            while enble == True:
                lasttrace = ''
                for x in range(0, lens):
                    if trace[x] in trace2[x]:
                        lasttrace = lasttrace + str(trace[x])
                    else:
                        lasttrace = lasttrace + str(n)
                if content.xpath(str(lasttrace).rstrip('/')):
                    text = content.xpath(str(lasttrace).rstrip('/') + '/@href')
                    if text == []:
                        text = ''
                    else:
                        text = text[0].strip()
                        if text[0] == '/':
                            text = str(host) + str(text)
                        elif text[0] == 'h':
                            text = str(text)
                        else:
                            text = str(host) + '/' + str(text)
                    info.append(text)
                    n += margin
                    enble = True
                else:
                    print('链接抓取中...')
                    enble = False
        except:
            print('抓取链接出错了')
            sys.exit()
    elif resultinfo[1] == 'result3':
        text = text[0].strip()
        if text[0] == '/':
            text = str(host) + str(text)
        elif text[0] == 'h':
            text = str(text)
        else:
            text = str(host) + '/' + str(text)
        info.append(text)
        print('链接抓取中...')
    else:
        print('抓取链接出错了')
    return info
    #正则获取信息判断
def judgezcontent(each, source):
    resultinfo = ['Z']
    source = source.get_text().strip()
    x1 = each[1]
    x2 = each[2]
    content = re.findall(str(x1) + '(.*?)' + str(x2), source, re.S)
    if content == '':
        print('无法获取到信息,可能输入有误')
        sys.exit()
    else:
        resultinfo.append(x1)
        resultinfo.append(x2)
    return resultinfo
    #利用正则获取数据或文字内容
def getzcontent(source, resultinfo):
    info = []
    source = source.get_text().strip()
    content = re.findall(str(resultinfo[1]) + '(.*?)' + str(resultinfo[2]), source, re.S)
    lencontent = len(content)
    for i in range(0, lencontent):
        if len(content[i]) < 20:
            info.append(content[i])
    print('信息抓取中...')
    return info
    #xpath判断图片下载
def judgeimg(each, content):
    result = ''
    resultinfo = ['M']
    trace = each[1]
    if not re.search('/img', str(trace)):
        print('该路径不可下载图片')
        sys.exit()
    text = content.xpath(str(trace) + '/@src')
    if text == []:
        trace = trace.replace('/tbody', '')
        text = content.xpath(str(trace) + '/@src')
    lentext = len(text)
    if not lentext == 1:
        if lentext == 0:
            print('无法下载图片')
            sys.exit()
        else:
            print('该处图片无法精确下载,下载后请根据需要筛选图片')
            result = result + 'result1'
            resultinfo.append(result)
            resultinfo.append(trace)
    else:
        text = content.xpath(str(trace) + '/@src')[0].strip()
        if re.search('javascript:', text):
            print('该图片无法下载')
            sys.exit()
        judge = ''
        if each[2] != '' and each[2] != None:
            judge = '是'
        if judge == '是':
            trace2 = each[2]
            if not re.search('/img', str(trace2)):
                print('该路径不可下载图片')
                sys.exit()
            text = content.xpath(str(trace2) + '/@src')
            if text == []:
                trace2 = trace2.replace('/tbody', '')
            while trace2 == trace:
                trace2 = pathagain()
                each[2] = trace2
                if not re.search('/img', str(trace2)):
                    print('该路径不可下载图片')
                    sys.exit()
                text = content.xpath(str(trace2) + '/@src')
                if text == []:
                    trace2 = trace2.replace('/tbody', '')
            if trace2 == '' or trace2 == None:
                result = result + 'result3'
                resultinfo.append(result)
                resultinfo.append(trace)
            else:
                result = result + 'result2'
                resultinfo.append(result)
                resultinfo.append(trace)
                resultinfo.append(trace2)
        else:
            text = text[0].strip()
            if re.search('javascript:', text):
                print('该图片无法下载')
                sys.exit()
            result = result + 'result3'
            resultinfo.append(result)
            resultinfo.append(trace)
    return resultinfo
    #利用xpath下载图片
def getimg(url, content, resultinfo):
    info = []
    protocol, s1 = urllib.parse.splittype(url)
    host, s2 = urllib.parse.splithost(s1)
    trace = resultinfo[2]
    text = content.xpath(str(trace) + '/@src')
    lentext = len(text)
    if resultinfo[1] == 'result1':
        for i in range(0, lentext):
            text = content.xpath(str(trace) + '/@src')
            if text[i][0] == '/':
                text = 'http://' + str(host) + str(text[i])
            elif text[i][0] == 'h':
                text = str(text[i])
            else:
                text = 'http://' + str(host) + '/' + str(text[i])
            info.append(text)
        print('图片下载中...')
    elif resultinfo[1] == 'result2':
        trace2 = resultinfo[3]
        enble = True
        num = []
        lens = len(trace)
        try:
            for x1 in range(0, lens):
                if trace[x1] not in trace2[x1]:
                    if trace[x1].isdigit() == True:
                        num.append(trace[x1])
            for x2 in range(0, lens):
                if trace2[x2] not in trace[x2]:
                    if trace2[x2].isdigit() == True:
                        num.append(trace2[x2])
            margin = int(num[1]) - int(num[0])
            n = int(num[0])
            while enble == True:
                lasttrace = ''
                for x in range(0, lens):
                    if trace[x] in trace2[x]:
                        lasttrace = lasttrace + str(trace[x])
                    else:
                        lasttrace = lasttrace + str(n)
                if content.xpath(str(lasttrace).rstrip('/')):
                    text = content.xpath(str(lasttrace).rstrip('/') + '/@src')
                    if text == []:
                        text = ''
                    else:
                        text = text[0].strip()
                        if text[0] == '/':
                            text = 'http://' + str(host) + str(text)
                        elif text[0] == 'h':
                            text = str(text)
                        else:
                            text = 'http://' + str(host) + '/' + str(text)
                    info.append(text)
                    n += margin
                    enble = True
                else:
                    print('图片下载中...')
                    enble = False
        except:
            print('下载出错了')
            sys.exit()
    elif resultinfo[1] == 'result3':
        text = text[0].strip()
        if text[0] == '/':
            text = 'http://' + str(host) + str(text)
        elif text[0] == 'h':
            text = str(text)
        else:
            text = 'http://' + str(host) + '/' + str(text)
        info.append(text)
        print('图片下载中...')
    else:
        print('下载出错了')
    return info
    #下载图片到本地
def downimg(info):
    imgpath = docpath + '\\图片下载\\'
    if os.path.exists(imgpath) == True:
        pass
    else:
        os.makedirs(imgpath)
    n = 1
    for each in info:
        try:
            timeinfo = time.strftime("%Y-%m-%d-%H-%M-%S-" + str(n))
            name = str(timeinfo)
            u = urllib.request.urlopen(each)
            img = u.read()
            f = open(imgpath + name + '.jpg', 'wb')
            f.write(img)
            f.close()
            n += 1
            time.sleep(0.5)
        except:
            continue
    return
    #根据信息类型选择相应的方式输出单页的信息
def getpageinfo(url, header, allresultinfo):
    pageinfo = []
    alen =len(allresultinfo)
    n = 1
    for each in allresultinfo:
        info = []
        html = getsource(url, header)
        if each[0] == 'Z':
            html = html.text
            content = BeautifulSoup(html, 'lxml')
            info = getzcontent(content, each)
        elif each[0] == 'U':
            html = html.content
            content = etree.HTML(html)
            info = getxurl(url, content, each)
        elif each[0] == 'X':
            html = html.content
            content = etree.HTML(html)
            info = getxcontent(content, each)
        elif each[0] == 'M':
            html = html.content
            content = etree.HTML(html)
            imginfo = getimg(url, content, each)
            downimg(imginfo)
        else:
            pass
        pageinfo.append(info)
        per = round(n/alen*100, 2)
        perc = str(per) + '%'
        n += 1
        print('已完成该页的', perc)
    return pageinfo
    #获取页面链接的方式1
def turnpage1(html, host, trace):
    html = html.content
    content = etree.HTML(html)
    pageurl = ''
    if not re.search('/a', str(trace)):
        print('按键路径错误')
        sys.exit()
    elif re.search('/a.*?/', str(trace)):
        mould = re.search('/a.*?/', str(trace)).group()
        place = trace.index(mould)
        trace = trace.replace(str(trace[place:]), str(mould))
        text = content.xpath(str(trace).rstrip('/') + '/@href')
        if text == []:
            trace = trace.replace('/tbody', '')
            text = content.xpath(str(trace) + '@href')
        if text == []:
            print('无法翻页,请使用自定义流程操作')
            sys.exit()
        else:
            text = content.xpath(str(trace) + '@href')[0].strip()
    elif re.search('/a/', str(trace)):
        place = trace.index('/a/')
        trace = trace.replace(str(trace[place:]), '/a')
        text = content.xpath(str(trace).rstrip('/') + '/@href')
        if text == []:
            trace = trace.replace('/tbody', '')
            text = content.xpath(str(trace) + '@href')
        if text == []:
            print('无法翻页,请使用自定义流程操作')
            sys.exit()
        else:
            text = content.xpath(str(trace) + '@href')[0].strip()
    else:
        text = content.xpath(str(trace) + '/@href')
        if text == []:
            trace = trace.replace('/tbody', '')
            text = content.xpath(str(trace) + '/@href')
        text = text[0].strip()
    if text[0] == '/':
        pageurl = pageurl + str(host) + str(text)
    elif text[0] == 'h':
        pageurl = pageurl + str(text)
    else:
        pageurl = pageurl + str(host) + '/' + str(text)
    return pageurl
    #获取页面链接的方式2
def turnpage2(html, host):
    html = html.text
    content = BeautifulSoup(html, 'lxml')
    urlinfo = content.find_all('a')
    pageurl = ''
    keys = ['下一页', '下一页>', '下页', '下页>', '后页', '后页>']
    for each in urlinfo:
        key = each.get_text().strip()
        text = each.get('href')
        if key in keys:
            if text[0] == '/':
                pageurl = pageurl + str(host) + str(text)
            elif text[0] == 'h':
                pageurl = pageurl + str(text)
            else:
                pageurl = pageurl + str(host) + '/' + str(text)
        else:
            pass
    return pageurl
    #综合获取页面链接的方式
def turnpage(url, html, trace):
    protocol, s1 = urllib.parse.splittype(url)
    host, s2 = urllib.parse.splithost(s1)
    pageurl = turnpage2(html, host)
    if pageurl == '':
        pageurl = turnpage1(html, host, trace)
    if 'http' not in pageurl:
        pageurl = 'http://' + pageurl
    return pageurl
    #获取所有页面的内容
def getallinfo1(url, header, auto):
    print('本次抓取可抓取页面内各类内容如文字/数据/链接/图片,若是无法抓取,请使用自定义流程尝试')
    nexturl = url
    allinfo = []
    enble = True
    pagenum = ''
    pagetrace = ''
    allresultinfo = []
    while enble == True:
        try:
            if auto[0] == '是':
                ways = infovalue3[1:]
                pagenum = infovalue3[0][11]
                pagetrace = infovalue3[0][10]
            else:
                ways = whichway()
                pagenum = ways[-1][1]
                pagetrace = ways[-1][0]
            allresultinfo = choseway(ways, url, header, auto)
            enble = False
        except:
            enble = True
    if pagenum != '' and pagenum != None:
        for i in range(0, int(pagenum)):
            try:
                html = getsource(nexturl, header)
                pageinfo = getpageinfo(url, header, allresultinfo)
                leninfo = len(pageinfo)
                lenallinfo = len(allinfo)
                if not lenallinfo == leninfo:
                    if leninfo > lenallinfo:
                        margin = leninfo - lenallinfo
                        for ii in range(0, margin):
                            allinfo.append([])
                    else:
                        print('在第' + str(i + 1) + '页抓取失败')
                        sys.exit()
                for iii in range(0, leninfo):
                    allinfo[iii] += pageinfo[iii]
                pageurl = turnpage(nexturl, html, pagetrace)
                if pageurl == nexturl:
                    print('在第' + str(i + 1) + '页翻页失败')
                    sys.exit()
                else:
                    print('第' + str(i + 1) + '页抓取完毕')
                    nexturl = quote(pageurl, safe='/:?=&;@$+,')
                time.sleep(1)
            except:
                continue
    else:
        allinfo = getpageinfo(url, header, allresultinfo)
    return allinfo

#抓取页面中各部分的详情页的内容
    #获取单页的每个详情页链接
def getvalue_detail():
    truevalue = []
    root = Tk()
    root.resizable(False, False)
    root.title("详情页设置")
    mainframe = ttk.Frame(root)
    mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
    screenwidth = root.winfo_screenwidth()
    screenheight = root.winfo_screenheight()
    root.geometry('370x270+%s+%s' % (screenwidth - 850, screenheight - 600))
    ttk.Label(mainframe, text='设置要抓取的详情页路径,通过该路径进到详情页抓取你想要的信息.').grid(column=0, columnspan=4, row=0, sticky=W)
    ttk.Label(mainframe, text='第二个详情页应是由第一个详情页顺序数(列或行)的第二个,勿跳序.').grid(column=0, columnspan=4, row=1, sticky=W)
    Label(mainframe, text='').grid(column=0, columnspan=4, row=2)
    ttk.Label(mainframe, text='第一个详情页的路径:').grid(column=0, row=3, sticky=W)
    var = StringVar()
    ttk.Entry(mainframe, width=30, textvariable=var).grid(column=1, row=3, sticky=W)
    ttk.Label(mainframe, text='第二个详情页的路径:').grid(column=0, row=4, sticky=W)
    var2 = StringVar()
    ttk.Entry(mainframe, width=30, textvariable=var2).grid(column=1, row=4, sticky=W)
    var3 = StringVar()
    var4 = StringVar()
    ttk.Label(mainframe, text='').grid(column=4, row=5)
    ttk.Label(mainframe, text='若需翻页请粘贴\n<下一页>按键的路径:').grid(column=0, row=6, sticky=W)
    ttk.Entry(mainframe, width=20, textvariable=var3).grid(column=1, row=6, sticky=W)
    ttk.Label(mainframe, text='需要抓取几页:').grid(column=0, row=7, sticky=W)
    ttk.Entry(mainframe, width=10, textvariable=var4).grid(column=1, row=7, sticky=W)
    ttk.Label(root, text='').grid(column=0, columnspan=4, row=8)
    ttk.Label(root, text='确认无误后请点击确定').grid(column=0, columnspan=4, row=9)
    ttk.Button(root, text='确定', command=lambda :root.destroy()).grid(column=0, columnspan=4, row=10)
    root.mainloop()
    x1 = var.get()
    x2 = var2.get()
    x3 = var3.get()
    x4 = var4.get()
    truevalue.append(x1)
    truevalue.append(x2)
    truevalue.append(x3)
    truevalue.append(x4)
    infovalue2[0][12] = x1
    infovalue2[0][13] = x2
    infovalue2[0][14] = x3
    infovalue2[0][15] = x4
    return truevalue
def judgedetailurl(detailinfo, html):
    html = html.content
    content = etree.HTML(html)
    result = ''
    resultinfo = ['DU']
    trace = detailinfo[0]
    if not re.search('/a', str(trace)):
        print('该内容不包含详情页链接')
        sys.exit()
    else:
        if re.search('/a.*?/', str(trace)):
            mould = re.search('/a.*?/', str(trace)).group()
            place = trace.index(mould)
            trace = trace.replace(str(trace[place:]), str(mould))
        elif re.search('/a/', str(trace)):
            place = trace.index('/a/')
            trace = trace.replace(str(trace[place:]), '/a')
    text = content.xpath(str(trace).rstrip('/') + '/@href')
    if text == []:
        trace = trace.replace('/tbody', '')
        text = content.xpath(str(trace) + '/@href')
    lentext = len(text)
    if not lentext == 1:
        if lentext == 0:
            print('无法获取详情页')
            sys.exit()
        else:
            print('详情页无法精确提取,有出错的可能性;若出错,可重启程序使用自定义流程操作')
            result = result + 'result1'
            resultinfo.append(result)
            resultinfo.append(trace)
    else:
        text = content.xpath(str(trace) + '/@href')[0].strip()
        if re.search('javascript:', text):
            print('详情页链接无法抓取')
            sys.exit()
        judge = ''
        if detailinfo[1] != '' and detailinfo[1] != None:
            judge = '是'
        if judge == '是':
            trace2 = detailinfo[1]
            if not re.search('/a', str(trace2)):
                print('该路径不包含详情页链接')
                sys.exit()
            else:
                if re.search('/a.*?/', str(trace2)):
                    mould = re.search('/a.*?/', str(trace2)).group()
                    place = trace2.index(mould)
                    trace2 = trace2.replace(str(trace2[place:]), str(mould))
                elif re.search('/a/', str(trace2)):
                    place = trace2.index('/a/')
                    trace2 = trace2.replace(str(trace2[place:]), '/a')
            text = content.xpath(str(trace2).rstrip('/') + '/@href')
            if text == []:
                trace2 = trace2.replace('/tbody', '')
            while trace2 == trace:
                print('两个详情页链接路径一样,无法获取多个链接,请重新输入')
                trace2 = pathagain()
                detailinfo[1] = trace2
                if not re.search('/a', str(trace2)):
                    print('该路径不包含详情页链接')
                    sys.exit()
                else:
                    if re.search('/a.*?/', str(trace2)):
                        mould = re.search('/a.*?/', str(trace2)).group()
                        place = trace2.index(mould)
                        trace2 = trace2.replace(str(trace2[place:]), str(mould))
                    elif re.search('/a/', str(trace2)):
                        place = trace2.index('/a/')
                        trace2 = trace2.replace(str(trace2[place:]), '/a')
                text = content.xpath(str(trace2).rstrip('/') + '/@href')
                if text == []:
                    trace2 = trace2.replace('/tbody', '')
            if trace2 == ''or trace2 == None:
                result = result + 'result3'
                resultinfo.append(result)
                resultinfo.append(trace)
            else:
                result = result + 'result2'
                resultinfo.append(result)
                resultinfo.append(trace)
                resultinfo.append(trace2)
        else:
            text = text[0].strip()
            if re.search('javascript:', text):
                print('详情页链接无法抓取')
                sys.exit()
            result = result + 'result3'
            resultinfo.append(result)
            resultinfo.append(trace)
    return resultinfo
def getdetailurl(url, content, resultinfo):
    info = []
    protocol, s1 = urllib.parse.splittype(url)
    host, s2 = urllib.parse.splithost(s1)
    trace = resultinfo[2]
    text = content.xpath(str(trace) + '/@href')
    lentext = len(text)
    if resultinfo[1] == 'result1':
        for i in range(0, lentext):
            text = content.xpath(str(trace) + '/@href')
            if text[i][0] == '/':
                text = str(host) + str(text[i])
            elif text[i][0] == 'h':
                text = str(text[i])
            else:
                text = str(host) + '/' + str(text[i])
            info.append(text)
        print('详情页链接抓取中...')
    elif resultinfo[1] == 'result2':
        trace2 = resultinfo[3]
        enble = True
        num = []
        lens = len(trace)
        try:
            for x1 in range(0, lens):
                if trace[x1] not in trace2[x1]:
                    if trace[x1].isdigit() == True:
                        num.append(trace[x1])
            for x2 in range(0, lens):
                if trace2[x2] not in trace[x2]:
                    if trace2[x2].isdigit() == True:
                        num.append(trace2[x2])
            margin = int(num[1]) - int(num[0])
            n = int(num[0])
            while enble == True:
                lasttrace = ''
                for x in range(0, lens):
                    if trace[x] in trace2[x]:
                        lasttrace = lasttrace + str(trace[x])
                    else:
                        lasttrace = lasttrace + str(n)
                if content.xpath(str(lasttrace).rstrip('/')):
                    text = content.xpath(str(lasttrace).rstrip('/') + '/@href')
                    if text == []:
                        text = ''
                    else:
                        text = text[0].strip()
                        if text[0] == '/':
                            text = str(host) + str(text)
                        elif text[0] == 'h':
                            text = str(text)
                        else:
                            text = str(host) + '/' + str(text)
                    info.append(text)
                    n += margin
                    enble = True
                else:
                    print('详情页链接抓取中...')
                    enble = False
        except:
            print('抓取出错了')
            sys.exit()
    elif resultinfo[1] == 'result3':
        text = text[0].strip()
        if text[0] == '/':
            text = str(host) + str(text)
        elif text[0] == 'h':
            text = str(text)
        else:
            text = str(host) + '/' + str(text)
        info.append(text)
        print('详情页链接抓取中...')
    else:
        print('抓取出错了')
    return info
    #获取详情页内的数据
    #获取单页的详情页内的总数据
def getpagedetailinfo(html, linkresultinfo, url, allresultinfo):
    html = html.content
    content = etree.HTML(html)
    linkinfo = getdetailurl(url, content, linkresultinfo)
    if 'http' not in linkinfo[0]:
        linkinfo[0] = 'http://' + linkinfo[0]
    pageallinfo = []
    llen = len(linkinfo)
    n = 1
    for each in linkinfo:
        per = '开始抓取第' + str(n) + '/' + str(llen) + '个详情页'
        n += 1
        print(per)
        if 'http' not in each:
            each = 'http://' + each
        try:
            pageinfo = getpageinfo(each, header, allresultinfo)
            leninfo = len(pageinfo)
            lenallinfo = len(pageallinfo)
            if not lenallinfo == leninfo:
                if leninfo > lenallinfo:
                    margin = leninfo - lenallinfo
                    for ii in range(0, margin):
                        pageallinfo.append([])
                else:
                    print('在该详情页' + str(each) + '抓取失败')
                    sys.exit()
            for iii in range(0, leninfo):
                pageallinfo[iii] += pageinfo[iii]
            time.sleep(1)
        except:
            continue
    return pageallinfo
    #获取所有页面内所有详情的信息
def getallinfo2(url, header, auto):
    print('本次抓取可抓取各个标题内详情页里的内容,若是无法抓取,请使用自定义流程尝试')
    html = getsource(url, header)
    enble1 = True
    linkresultinfo= []
    detailinfo = []
    while enble1 == True:
        try:
            if auto[0] == '是':
                detailinfo = infovalue3[0][12:16]
            else:
                detailinfo = getvalue_detail()
            linkresultinfo = judgedetailurl(detailinfo, html)
            enble1 = False
        except:
            enble1 = True
    html = html.content
    content = etree.HTML(html)
    linkinfo = getdetailurl(url, content, linkresultinfo)
    if 'http' not in linkinfo[0]:
        linkinfo[0] = 'http://' + linkinfo[0]
    enble = True
    allresultinfo = []
    print('添加数据的操作请到详情页内进行:')
    while enble == True:
        try:
            if auto[0] == '是':
                ways = infovalue3[1:]
            else:
                ways = whichway()
            allresultinfo = choseway(ways, linkinfo[0], header, auto)
            enble = False
        except:
            enble = True
    pagenum = detailinfo[3]
    pagetrace = detailinfo[2]
    nexturl = url
    allinfo = []
    if pagenum != '' and pagenum != None:
        for i in range(0, int(pagenum)):
            try:
                html = getsource(nexturl, header)
                pageallinfo = getpagedetailinfo(html, linkresultinfo, url, allresultinfo)
                leninfo = len(pageallinfo)
                lenallinfo = len(allinfo)
                if not lenallinfo == leninfo:
                    if leninfo > lenallinfo:
                        margin = leninfo - lenallinfo
                        for ii in range(0, margin):
                            allinfo.append([])
                    else:
                        print('在第' + str(i + 1) + '页抓取失败')
                        sys.exit()
                for iii in range(0, leninfo):
                    allinfo[iii] += pageallinfo[iii]
                pageurl = turnpage(nexturl, html, pagetrace)
                if pageurl == nexturl:
                    print('在第' + str(i + 1) + '页翻页失败')
                    sys.exit()
                else:
                    print('第' + str(i + 1) + '页抓取完毕')
                    nexturl = quote(pageurl, safe='/:?=&;@$+,')
                time.sleep(1)
            except:
                continue
    else:
        html = getsource(nexturl, header)
        allinfo = getpagedetailinfo(html, linkresultinfo, url, allresultinfo)
    return allinfo

#浏览器模拟抓取
    #Chrome模拟翻页
def chromepage():
    list = []
    root = Tk()
    root.resizable(False, False)
    root.title("翻页设置")
    mainframe = ttk.Frame(root)
    mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
    screenwidth = root.winfo_screenwidth()
    screenheight = root.winfo_screenheight()
    root.geometry('250x160+%s+%s' % (screenwidth - 850, screenheight - 600))
    ttk.Label(mainframe, text="若无需翻页可直接点击'确定'跳过").grid(column=0, columnspan=4, row=0, sticky=W)
    var = StringVar()
    var2 = StringVar()
    ttk.Label(mainframe, text='').grid(column=4, row=5)
    ttk.Label(mainframe, text='若需翻页请粘贴\n<下一页>按键的路径:').grid(column=0, row=6, sticky=W)
    ttk.Entry(mainframe, width=20, textvariable=var).grid(column=1, row=6, sticky=W)
    ttk.Label(mainframe, text='需要抓取几页:').grid(column=0, row=7, sticky=W)
    ttk.Entry(mainframe, width=10, textvariable=var2).grid(column=1, row=7, sticky=W)
    ttk.Label(root, text='').grid(column=0, columnspan=4, row=8)
    ttk.Label(root, text='确认无误后请点击确定').grid(column=0, columnspan=4, row=9)
    ttk.Button(root, text='确定', command=lambda: root.destroy()).grid(column=0, columnspan=4, row=10)
    root.mainloop()
    x1 = var.get()
    x2 = var2.get()
    list.append(x1)
    list.append(x2)
    return list
def chrometurnpage1(drive):
    keys = ['下一页', '下一页>', '下页', '下页>', '后页', '后页>']
    enble = True
    while enble == True:
        for key in keys:
            try:
                if drive.find_element_by_link_text(key):
                    drive.find_element_by_link_text(key).click()
                    enble = False
            except:
                continue
def chrometurnpage2(drive, trace):
    if not re.search('/a', str(trace)):
        print('按键路径错误')
        sys.exit()
    elif re.search('/a.*?/', str(trace)):
        mould = re.search('/a.*?/', str(trace)).group()
        place = trace.index(mould)
        trace = trace.replace(str(trace[place:]), str(mould))
        trace = trace.rstrip('/')
    elif re.search('/a/', str(trace)):
        place = trace.index('/a/')
        trace = trace.replace(str(trace[place:]), '/a')
    drive.find_element_by_xpath(trace).click()
def chrometurnpage(drive, trace):
    try:
        chrometurnpage1(drive)
    except:
        chrometurnpage2(drive, trace)
    #选择下一步流程
def help1():
    root = Tk()
    root.resizable(False, False)
    root.title("注意事项")
    mainframe = ttk.Frame(root)
    mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
    screenwidth = root.winfo_screenwidth()
    screenheight = root.winfo_screenheight()
    root.geometry('400x200+%s+%s' % (screenwidth - 650, screenheight - 500))
    ttk.Label(root, text='1.点击项-可在下一步点击你粘贴了路径的按键;若要抓取详情页,则粘贴两').grid(row=0, sticky=W)
    ttk.Label(root, text='个同列详情页点击进入的路径就会在执行后依次点开了;').grid(row=1, sticky=W)
    ttk.Label(root, text='2.传入项-可传入值给粘贴了路径的输入框,可用于搜索栏处;').grid(row=2, sticky=W)
    ttk.Label(root, text='3.获取项-可获取粘贴了路径处的文本/数据/链接/图片,和点击一样,只要').grid(row=3, sticky=W)
    ttk.Label(root, text='粘贴同列的两个内容就会在执行后依次抓取了;').grid(row=4, sticky=W)
    ttk.Label(root, text='4.下拉项-表示鼠标可下拉到最底部.适用于需要下拉才加载内容的页面,').grid(row=5, sticky=W)
    ttk.Label(root, text='建议拉到抓取的地方时再开始执行;').grid(row=6, sticky=W)
    ttk.Label(root, text='5.点击执行后会根据选择的操作依次进行,即从开始抓取的页面模拟一次抓').grid(row=7, sticky=W)
    ttk.Label(root, text='取信息的操作流程即可.').grid(row=8, sticky=W)
    ttk.Button(root, text='确认', command=lambda: root.destroy()).grid(row=9)
    root.mainloop()
def chose1(root, list):
    list.append('C')
    root.destroy()
def chose2(root, list):
    list.append('S')
    root.destroy()
def chose3(root, list):
    list.append('G')
    root.destroy()
def chose4(root, list):
    list.append('F')
    root.destroy()
def chromechose():
    list = []
    root = Tk()
    root.resizable(False, False)
    root.title("流程设置")
    mainframe = ttk.Frame(root)
    mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
    screenwidth = root.winfo_screenwidth()
    screenheight = root.winfo_screenheight()
    root.geometry('220x170+%s+%s' % (screenwidth - 850, screenheight - 600))
    ttk.Label(mainframe, text='        ').grid(column=0, rowspan=7)
    ttk.Label(mainframe, text="请选择下一步的操作").grid(column=1, columnspan=4, row=0)
    ttk.Button(mainframe, text='注意事项', command=lambda: help1()).grid(column=1, columnspan=4, row=2)
    ttk.Label(mainframe, text='').grid(columnspan=4, row=3)
    ttk.Button(mainframe, text='点击', command=lambda: chose1(root, list)).grid(column=1, row=4)
    ttk.Label(mainframe, text='     ').grid(column=2, row=4)
    ttk.Button(mainframe, text='传入', command=lambda: chose2(root, list)).grid(column=3, row=4)
    ttk.Button(mainframe, text='获取', command=lambda: chose3(root, list)).grid(column=1, row=5)
    ttk.Label(mainframe, text='     ').grid(column=2, row=5)
    ttk.Button(mainframe, text='下拉', command=lambda: chose4(root, list)).grid(column=3, row=5)
    ttk.Label(mainframe, text='').grid(columnspan=4, row=6)
    ttk.Button(mainframe, text='开始执行', command=lambda: root.destroy()).grid(column=1, columnspan=4, row=7)
    root.mainloop()
    return list
    #判断下一步的具体操作
def chromesend():
    list = []
    root = Tk()
    root.resizable(False, False)
    root.title("传入项")
    mainframe = ttk.Frame(root)
    mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
    screenwidth = root.winfo_screenwidth()
    screenheight = root.winfo_screenheight()
    root.geometry('250x140+%s+%s' % (screenwidth - 850, screenheight - 600))
    var = StringVar()
    var2 = StringVar()
    ttk.Label(mainframe, text='').grid(column=4, row=5)
    ttk.Label(mainframe, text='请粘贴输入框的路径:').grid(column=0, row=6, sticky=W)
    ttk.Entry(mainframe, width=20, textvariable=var).grid(column=1, row=6, sticky=W)
    ttk.Label(mainframe, text='请输入要传入的值:').grid(column=0, row=7, sticky=W)
    ttk.Entry(mainframe, width=10, textvariable=var2).grid(column=1, row=7, sticky=W)
    ttk.Label(root, text='').grid(column=0, columnspan=4, row=8)
    ttk.Label(root, text='确认无误后请点击确定').grid(column=0, columnspan=4, row=9)
    ttk.Button(root, text='确定', command=lambda: root.destroy()).grid(column=0, columnspan=4, row=10)
    root.mainloop()
    x1 = var.get()
    x2 = var2.get()
    list.append(x1)
    list.append(x2)
    return list
def chromeclick():
    list = []
    root = Tk()
    root.resizable(False, False)
    root.title("点击项")
    mainframe = ttk.Frame(root)
    mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
    screenwidth = root.winfo_screenwidth()
    screenheight = root.winfo_screenheight()
    root.geometry('280x140+%s+%s' % (screenwidth - 850, screenheight - 600))
    var = StringVar()
    var2 = StringVar()
    ttk.Label(mainframe, text='').grid(column=4, row=5)
    ttk.Label(mainframe, text='请粘贴要点击的路径1:').grid(column=0, row=6, sticky=W)
    ttk.Entry(mainframe, width=20, textvariable=var).grid(column=1, row=6, sticky=W)
    ttk.Label(mainframe, text='请粘贴要点击的路径2:').grid(column=0, row=7, sticky=W)
    ttk.Entry(mainframe, width=20, textvariable=var2).grid(column=1, row=7, sticky=W)
    ttk.Label(root, text='').grid(column=0, columnspan=4, row=8)
    ttk.Label(root, text='确认无误后请点击确定').grid(column=0, columnspan=4, row=9)
    ttk.Button(root, text='确定', command=lambda: root.destroy()).grid(column=0, columnspan=4, row=10)
    root.mainloop()
    x1 = var.get()
    x2 = var2.get()
    list.append(x1)
    list.append(x2)
    return list
def chromelocal():
    root = Tk()
    root.resizable(False, False)
    root.title("定位")
    mainframe = ttk.Frame(root)
    mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
    screenwidth = root.winfo_screenwidth()
    screenheight = root.winfo_screenheight()
    root.geometry('280x140+%s+%s' % (screenwidth - 850, screenheight - 600))
    var = StringVar()
    ttk.Label(mainframe, text='该路径有多个内容,可通过路径处的完整文本信息来\n'
                              '精确定位点击的地方,若无需定位则直接按确定即可').grid(column=0, columnspan=4, row=3)
    ttk.Label(mainframe, text='').grid(column=0, columnspan=4, row=5)
    ttk.Label(mainframe, text='路径处的完整文本信息:').grid(column=0, row=6, sticky=W)
    ttk.Entry(mainframe, width=20, textvariable=var).grid(column=1, row=6, sticky=W)
    ttk.Label(mainframe, text='').grid(column=0, columnspan=4, row=8)
    ttk.Label(mainframe, text='确认无误后请点击确定').grid(column=0, columnspan=4, row=9)
    ttk.Button(mainframe, text='确定', command=lambda: root.destroy()).grid(column=0, columnspan=4, row=10)
    root.mainloop()
    x = var.get()
    return x
def get1(root, list):
    list.append('X')
    root.destroy()
def get2(root, list):
    list.append('U')
    root.destroy()
def get3(root, list):
    list.append('M')
    root.destroy()
def chromeget():
    list = []
    root = Tk()
    root.resizable(False, False)
    root.title("获取项")
    mainframe = ttk.Frame(root)
    mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
    screenwidth = root.winfo_screenwidth()
    screenheight = root.winfo_screenheight()
    root.geometry('350x140+%s+%s' % (screenwidth - 850, screenheight - 600))
    var = StringVar()
    var2 = StringVar()
    ttk.Label(mainframe, text='').grid(column=4, row=5)
    ttk.Label(mainframe, text='请粘贴要获取内容的路径1:').grid(column=0, row=6, sticky=W)
    ttk.Entry(mainframe, width=25, textvariable=var).grid(column=1, columnspan=3, row=6, sticky=W)
    ttk.Label(mainframe, text='请粘贴要获取内容的路径2:').grid(column=0, row=7, sticky=W)
    ttk.Entry(mainframe, width=25, textvariable=var2).grid(column=1, columnspan=3, row=7, sticky=W)
    ttk.Label(mainframe, text='').grid(column=0, columnspan=4, row=8)
    ttk.Button(mainframe, text='文本/数据', command=lambda: get1(root, list)).grid(column=0, row=10)
    ttk.Button(mainframe, text='链接', command=lambda: get2(root, list)).grid(column=1, row=10)
    ttk.Label(mainframe, text='          ').grid(column=2, row=10)
    ttk.Button(mainframe, text='图片', command=lambda: get3(root, list)).grid(column=3, row=10)
    root.mainloop()
    x1 = var.get()
    x2 = var2.get()
    list.append(x1)
    list.append(x2)
    return list
def chromewhich(drive):
    enble = True
    while enble == True:
        list = chromechose()
        try:
            if list == []:
                enble = False
            elif list[0] == 'C':
                alist = chromeclick()
                s1 = drive.find_elements_by_xpath(alist[0])
                if s1 == []:
                    print('路径出错,无法获取')
                    sys.exit()
                if not len(s1) == 1:
                    print('同路径有多个内容,点击进入的网页可能不精确,可自行打开需要抓取的页面再操作;若是结构相同的网页,可照常操作')
                    if alist[1] == '':
                        x = chromelocal()
                        alist.append(x)
                        if not alist[2] == '':
                            s1 = drive.find_elements_by_link_text(alist[2])
                    alist[1] = ''
                if alist[0] == alist[1]:
                    print('两个路径一样,请重新输入')
                    sys.exit()
                for each in alist:
                    list.append(each)
                infovalue2.append(list)
                s1[0].click()
                handles = drive.window_handles
                drive.switch_to.window(handles[-1])
                enble = True
            elif list[0] == 'S':
                clist = chromesend()
                s2 = drive.find_elements_by_xpath(clist[0])
                if s2 == []:
                    print('路径出错,无法获取')
                    sys.exit()
                for each in clist:
                    list.append(each)
                infovalue2.append(list)
                drive.find_element_by_xpath(clist[0]).clear()
                drive.find_element_by_xpath(clist[0]).send_keys(clist[1])
                enble = True
            elif list[0] == 'G':
                blist = chromeget()
                s3 = drive.find_elements_by_xpath(blist[1])
                if s3 == []:
                    print('路径出错,无法获取')
                    sys.exit()
                if not len(s3) == 1:
                    print('同路径有多个内容,抓取的内容可能不精确,请照常抓取后根据需要筛选')
                    blist[2] = ''
                if blist[1] == blist[2]:
                    print('两个路径一样,请重新输入')
                    sys.exit()
                infovalue2.append(blist)
                enble = True
            elif list[0] == 'F':
                infovalue2.append(list)
                drive.execute_script("window.scrollBy(0,5000)")
                enble = True
            else:
                enble = True
        except:
            print('路径出错,请勿修改并重新输入')
            continue
    #根据获取的信息执行操作
def chromegetcontent1(each, drive):
    info = []
    trace = each[1]
    trace2 = each[2]
    enble = True
    num = []
    lens = len(trace)
    try:
        for x1 in range(0, lens):
            if trace[x1] not in trace2[x1]:
                if trace[x1].isdigit() == True:
                    num.append(trace[x1])
        for x2 in range(0, lens):
            if trace2[x2] not in trace[x2]:
                if trace2[x2].isdigit() == True:
                    num.append(trace2[x2])
        margin = int(num[1]) - int(num[0])
        n = int(num[0])
        while enble == True:
            lasttrace = ''
            for x in range(0, lens):
                if trace[x] in trace2[x]:
                    lasttrace = lasttrace + str(trace[x])
                else:
                    lasttrace = lasttrace + str(n)
            if not drive.find_elements_by_xpath(str(lasttrace).rstrip('/')) == []:
                text = ''
                if each[0] == 'X':
                    text = drive.find_element_by_xpath(str(lasttrace).rstrip('/')).text
                elif each[0] == 'U':
                    text = drive.find_element_by_xpath(str(lasttrace).rstrip('/')).get_attribute('href')
                elif each[0] == 'M':
                    text = drive.find_element_by_xpath(str(lasttrace).rstrip('/')).get_attribute('src')
                elif each[0] == 'C':
                    text = drive.find_element_by_xpath(str(lasttrace).rstrip('/'))
                info.append(text)
                n += margin
                enble = True
            else:
                print('信息抓取中...')
                enble = False
        if each[0] == 'M':
            downimg(info)
    except:
        print('抓取信息出错了')
        sys.exit()
    return info
def chromegetcontent2(each, drive):
    info = []
    xxx = drive.find_elements_by_xpath(each[1])
    lens = len(xxx)
    text = ''
    if lens == 1:
        if each[0] == 'X':
            text = drive.find_element_by_xpath(each[1]).text
        elif each[0] == 'U':
            text = drive.find_element_by_xpath(each[1]).get_attribute('href')
        elif each[0] == 'M':
            text = drive.find_element_by_xpath(each[1]).get_attribute('src')
        info.append(text)
    else:
        for i in xxx:
            if each[0] == 'X':
                text = i.text
            elif each[0] == 'U':
                text = i.get_attribute('href')
            elif each[0] == 'M':
                text = i.get_attribute('src')
            info.append(text)
    if each[0] == 'M':
        downimg(info)
    return info
def chromestart(drive):
    startlist = infovalue2
    pageallinfo = []
    clist = []
    n = 0
    for each in startlist:
        if each[0] == 'C' and each[2] != '':
            n = startlist.index(each)
        elif each[0] == 'C' and each[2] == '':
            lens = drive.find_elements_by_xpath(each[1])
            if not lens == 1:
                n = startlist.index(each)
        clist = startlist[n]
    startlist = infovalue2[1:]
    if n == 0:
        pageinfo = []
        for each in startlist:
            i = each[0]
            if i == 'C':
                if len(each) == 4 and each[3] != '':
                    drive.find_element_by_link_text(each[3]).click()
                else:
                    drive.find_element_by_xpath(each[1]).click()
                time.sleep(1)
                handles = drive.window_handles
                drive.switch_to.window(handles[-1])
            elif i == 'S':
                drive.find_element_by_xpath(each[1]).clear()
                drive.find_element_by_xpath(each[1]).send_keys(each[2])
            elif i == 'F':
                drive.execute_script("window.scrollBy(0,5000)")
            else:
                if not each[2] == '':
                    info = chromegetcontent1(each, drive)
                else:
                    info = chromegetcontent2(each, drive)
                pageinfo.append(info)
        pageallinfo = pageinfo
    else:
        pageinfo1 = []
        if not n-1 == 0:
            for each in startlist[:n - 1]:
                i = each[0]
                if i == 'C':
                    if len(each) == 4 and each[3] != '':
                        drive.find_element_by_link_text(each[3]).click()
                    else:
                        drive.find_element_by_xpath(each[1]).click()
                    time.sleep(1)
                    handles = drive.window_handles
                    drive.switch_to.window(handles[-1])
                elif i == 'S':
                    drive.find_element_by_xpath(each[1]).clear()
                    drive.find_element_by_xpath(each[1]).send_keys(each[2])
                elif i == 'F':
                    drive.execute_script("window.scrollBy(0,5000)")
                else:
                    if not each[2] == '':
                        info = chromegetcontent1(each, drive)
                    else:
                        info = chromegetcontent2(each, drive)
                    pageinfo1.append(info)
        if not clist[2] == '':
            clickinfo = chromegetcontent1(clist, drive)
        else:
            clickinfo = drive.find_elements_by_xpath(clist[1])
        for each in clickinfo:
            each.click()
            time.sleep(1)
            handles = drive.window_handles
            drive.switch_to.window(handles[-1])
            pageinfo = []
            for each2 in startlist[n:]:
                i = each2[0]
                if i == 'C':
                    if len(each) == 4 and each[3] != '':
                        drive.find_element_by_link_text(each[3]).click()
                    else:
                        drive.find_element_by_xpath(each2[1]).click()
                    time.sleep(1)
                    handles = drive.window_handles
                    drive.switch_to.window(handles[-1])
                elif i == 'S':
                    drive.find_element_by_xpath(each2[1]).clear()
                    drive.find_element_by_xpath(each2[1]).send_keys(each2[2])
                elif i == 'F':
                    drive.execute_script("window.scrollBy(0,5000)")
                else:
                    if not each2[2] == '':
                        info = chromegetcontent1(each2, drive)
                    else:
                        info = chromegetcontent2(each2, drive)
                    pageinfo.append(info)
            handles = drive.window_handles
            while len(handles) != 1:
                drive.switch_to.window(handles[-1])
                drive.close()
                handles = drive.window_handles
            drive.switch_to.window(handles[0])
            leninfo = len(pageinfo)
            lenallinfo = len(pageallinfo)
            if not lenallinfo == leninfo:
                if leninfo > lenallinfo:
                    margin = leninfo - lenallinfo
                    for ii in range(0, margin):
                        pageallinfo.append([])
                else:
                    print('在该详情页' + str(each) + '抓取失败')
                    sys.exit()
            for iii in range(0, leninfo):
                pageallinfo[iii] += pageinfo[iii]
            print('erroepag3', pageallinfo)
        if not pageinfo1 == []:
            pageallinfo.append(pageinfo1)
    return pageallinfo
    #获取所有页面的所有信息
def getallinfo3(url, auto):
    drive = webdriver.Chrome()
    try:
        drive.get(url)
        drive.maximize_window()
    except:
        print('抓取出错,请等待页面加载完毕再操作路径,请重试')
        sys.exit()
    if auto[0] == '是':
        lens = len(infovalue3)
        for i in range(0, lens-1):
            infovalue2.append([])
        for i in range(0, lens):
            infovalue2[i] = infovalue3[i]
    else:
        pagelist = chromepage()
        infovalue2[0][10] = pagelist[0]
        infovalue2[0][11] = pagelist[1]
        chromewhich(drive)
        drive.quit()
        drive = webdriver.Chrome()
        drive.get(url)
        drive.maximize_window()
    pagetrace = infovalue2[0][10]
    pagenum = infovalue2[0][11]
    allinfo = []
    if pagenum != '' and pagenum != None:
        for i in range(0, int(pagenum)):
            try:
                pageurl = drive.current_url
                pageallinfo = chromestart(drive)
                leninfo = len(pageallinfo)
                lenallinfo = len(allinfo)
                if not lenallinfo == leninfo:
                    if leninfo > lenallinfo:
                        margin = leninfo - lenallinfo
                        for ii in range(0, margin):
                            allinfo.append([])
                    else:
                        print('在第' + str(i + 1) + '页抓取失败')
                        sys.exit()
                for iii in range(0, leninfo):
                    allinfo[iii] += pageallinfo[iii]
                chrometurnpage(drive, pagetrace)
                nexturl = drive.current_url
                if nexturl == pageurl:
                    print('在第' + str(i + 1) + '页翻页失败')
                    sys.exit()
                else:
                    print('第' + str(i + 1) + '页抓取完毕')
                time.sleep(1)
            except:
                continue
    else:
        allinfo = chromestart(drive)
    drive.quit()
    return allinfo

#保存数据
    #保存数据到excel
def get_save1(root, list):
    list.append('E')
    root.destroy()
def get_save2(root, list):
    list.append('T')
    root.destroy()
def getvalue_save():
    truevalue = []
    root = Tk()
    root.resizable(False, False)
    root.title("保存设置")
    mainframe = ttk.Frame(root)
    mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
    screenwidth = root.winfo_screenwidth()
    screenheight = root.winfo_screenheight()
    root.geometry('270x100+%s+%s' % (screenwidth - 850, screenheight - 600))
    ttk.Label(mainframe, text='    ').grid(column=0, row=0)
    ttk.Label(mainframe, text='请输入保存数据的文件名,点击导出即保存').grid(column=1, columnspan=2, row=0)
    var = StringVar()
    ttk.Entry(mainframe, width=30, textvariable=var).grid(column=1, columnspan=2, row=1)
    ttk.Label(mainframe, text='    ').grid(column=0, row=2)
    ttk.Button(mainframe, text='导出Excel', command=lambda: get_save1(root, truevalue)).grid(column=1, row=3)
    ttk.Button(mainframe, text='导出Txt', command=lambda: get_save2(root, truevalue)).grid(column=2, row=3)
    root.mainloop()
    x = var.get()
    truevalue.append(x)
    return truevalue
def saveexcel(allinfo, docname):
    blen = len(allinfo)
    doc = docname + '.xlsx'
    path = docpath + '\\' + infovalue1[0] + '\\' + doc
    if os.path.exists(path) == True:
        print('该文档已存在,请重新取名')
        sys.exit()
    else:
        if infovalue1[0] != '' and infovalue1[0] != None:
            actpath = docpath + '\\' + infovalue1[0]
            if not os.path.exists(actpath) == True:
                os.makedirs(actpath)
        workbook = xlsxwriter.Workbook(path)
        workbook.close()
        openpath = openpyxl.load_workbook(path)
        sheet = openpath.get_sheet_names()
        ws = openpath.get_sheet_by_name(sheet[0])
        for i in range(0, blen):
            slen = len(allinfo[i])
            for ii in range(0, slen):
                ws.cell(row=ii+1, column=i+1, value=allinfo[i][ii] )
        openpath.save(path)
    return
    #保存数据到txt
def savetxt(allinfo, docname):
    blen = len(allinfo)
    doc = docname + '.txt'
    path = docpath + '\\' + infovalue1[0] + '\\' + doc
    if os.path.exists(path) == True:
        print('该文档已存在,请重新取名')
        sys.exit()
    else:
        if infovalue1[0] != '' and infovalue1[0] != None:
            actpath = docpath + '\\' + infovalue1[0]
            if not os.path.exists(actpath) == True:
                os.makedirs(actpath)
        f = open(path, 'w')
        for i in range(0, blen):
            slen = len(allinfo[i])
            for ii in range(0, slen):
                f.writelines(allinfo[i][ii] + '\n')
        f.close()
    return
    #选择保存数据的方式
def saveways(allinfo):
    list = ['\\', '/', ':', '*', '?', '"', '<', '>', '|']
    enble = True
    while enble == True:
        try:
            info = getvalue_save()
            docname = info[1]
            infovalue1[2] = docname
            if docname == '':
                sys.exit()
            for each in docname:
                for un in list:
                    if un in each:
                        print('名字存在非法字符,请重新设置')
                        sys.exit()
            if info[0] == 'E':
                saveexcel(allinfo, docname)
                tt = '是否需要以其他方式导出?'
                xx = getvalue_yandn(tt)
                choice = xx[0]
                if choice == '是':
                    sys.exit()
            elif info[0] == 'T':
                savetxt(allinfo, docname)
                tt = '是否需要以其他方式导出?'
                xx = getvalue_yandn(tt)
                choice = xx[0]
                if choice == '是':
                    sys.exit()
            else:
                print('导出方式输入错误,请重新输入')
                sys.exit()
            enble = False
        except:
            enble = True
    print('数据文件已保存至' + str(docpath) + '\\' + str(infovalue1[0]))
    print('保存完毕')
    return
def saveappend(allinfo):
    blen = len(allinfo)
    docname = infovalue1[2]
    doc1 = docname + '.txt'
    path1 = docpath + '\\' + infovalue1[0] + '\\' + doc1
    doc2 = docname + '.xlsx'
    path2 = docpath + '\\' + infovalue1[0] + '\\' + doc2
    if os.path.exists(path1) == True:
        f = open(path1, 'a')
        for i in range(0, blen):
            slen = len(allinfo[i])
            for ii in range(0, slen):
                f.writelines(allinfo[i][ii] + '\n')
        f.close()
    if os.path.exists(path2) == True:
        openpath = openpyxl.load_workbook(path2)
        sheet = openpath.get_sheet_names()
        ws = openpath.get_sheet_by_name(sheet[0])
        rm = ws.max_row
        for i in range(0, blen):
            slen = len(allinfo[i])
            for ii in range(0, slen):
                ws.cell(row=ii + 1 + rm, column=i + 1, value=allinfo[i][ii])
        openpath.save(path2)
    print('更新完毕')
    return
    #保存项目名称
def saveitem(info):
    openpath = openpyxl.load_workbook(doctxt)
    sheet = openpath.get_sheet_names()
    ws = openpath.get_sheet_by_name(sheet[1])
    ws.append(info)
    openpath.save(doctxt)

#抓取方式的选择
def choose1(root, list):
    list.append('1')
    root.destroy()
def choose2(root, list):
    list.append('2')
    root.destroy()
def choose3(root, list):
    list.append('3')
    root.destroy()
def getvalue_choose():
    truevalue = []
    root = Tk()
    root.resizable(False, False)
    root.title("抓取设置")
    mainframe = ttk.Frame(root)
    mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
    screenwidth = root.winfo_screenwidth()
    screenheight = root.winfo_screenheight()
    root.geometry('450x200+%s+%s' % (screenwidth - 850, screenheight - 600))
    Label(mainframe, text='').grid(column=0, columnspan=3, row=0)
    Label(mainframe, text='     ').grid(column=0, rowspan=8)
    ttk.Label(mainframe, text='请输入要抓取的初始网页的链接:').grid(column=1, columnspan=3, row=1, sticky=W)
    ttk.Button(mainframe, text='第1种方式', command=lambda: root.destroy()).grid(column=1, row=7)
    var = StringVar()
    if not infovalue3 == []:
        var.set(infovalue3[0][0])
    ttk.Entry(mainframe, width=30, textvariable=var).grid(column=1, columnspan=3, row=2, sticky=W)
    ttk.Label(mainframe, text='你可以选择以下三种方式抓取,点击即确认:').grid(column=1, columnspan=3, row=3, sticky=W)
    ttk.Label(mainframe, text='1,可翻页抓取页面上的信息如文字/数据/链接/图片;').grid(column=1, columnspan=3, row=4, sticky=W)
    ttk.Label(mainframe, text='2,可翻页抓取页面中每个模块的详情页内的上述各类信息;').grid(column=1, columnspan=3, row=5, sticky=W)
    ttk.Label(mainframe, text='3,可自定义抓取上述各类信息的步骤,适用于以上方式无法抓取到信息的网页.').grid(column=1, columnspan=3, row=6, sticky=W)
    ttk.Button(mainframe, text='第1种方式', command=lambda: choose1(root, truevalue)).grid(column=1, row=7)
    ttk.Button(mainframe, text='第2种方式', command=lambda: choose2(root, truevalue)).grid(column=2, row=7)
    ttk.Button(mainframe, text='第3种方式', command=lambda: choose3(root, truevalue)).grid(column=3, row=7)
    root.mainloop()
    x = var.get()
    truevalue.append(x)
    infovalue2[0][0] = truevalue[1]
    return truevalue

#主运行程序
def mainpro():
    enble = True
    while enble == True:
        theheader = header
        allinfo = []
        auto = ['']
        newer()
        judgeload = loadaction()
        if judgeload == '是':
            tt = '是否按照保存的操作记录进行自动更新?'
            xx = getvalue_yandn(tt)
            auto[0] = xx[0]
        iflogin = ''
        if auto[0] == '是':
            openpath = openpyxl.load_workbook(doctxt)
            sheet = openpath.get_sheet_names()
            ws = openpath.get_sheet_by_name(sheet[1])
            rm = ws.max_row
            which = ['', '']
            for i in range(0, rm):
                val = ws.cell(row=i+1, column=1).value
                if val == infovalue1[0]:
                    which[1] = infovalue3[0][0]
                    which[0] = ws.cell(row=i+1, column=2).value
                    infovalue1[2] = ws.cell(row=i+1, column=3).value
                    break
            if infovalue3[0][1] != None and infovalue3[0][1] != '':
                iflogin = '是' + iflogin
            print(iflogin)
        else:
            print('请输入真实网址,即粘贴在浏览器上不会跳转到其他网页的网址,否则无法抓取信息')
            which = getvalue_choose()
            while which[1] == '':
                which = getvalue_choose()
            infovalue1[1] = which[0]
            tt = '将要抓取的网页是否需要登录?'
            xx = getvalue_yandn(tt)
            iflogin = xx[0]
        if iflogin == '是':
            if not judgeload == '是':
                print('可加载过去的项目自动登录')
                judgeload = loadaction()
            theheader = postheader(judgeload)
            print('若是无法抓取到内容,请尝试重启程序重新登录')
        aimurl = which[1]
        aimurl = str(aimurl).replace(' ', '')
        if 'http' not in aimurl:
            aimurl = 'http://' + aimurl
        enble2 = True
        while enble2 == True:
            if which[0] == '1':
                allinfo = getallinfo1(aimurl, theheader, auto)
            elif which[0] == '2':
                allinfo = getallinfo2(aimurl, theheader, auto)
            elif which[0] == '3':
                allinfo = getallinfo3(aimurl, auto)
            if allinfo == []:
                print('什么也没有抓取到')
                if iflogin == '是':
                    tt = '登录可能过期,是否需要重新登录尝试再次抓取?'
                    xx = getvalue_yandn(tt)
                    relogin = xx[0]
                    if relogin == '是':
                        postcookie = []
                        cookie = savecookie(judgeload)
                        postcookie.append(cookie)
                        header['Cookie'] = postcookie[0]
                        enble2 = True
                    else:
                        enble2 = False
                else:
                    enble2 = False
            else:
                if not auto[0] == '是':
                    if judgeload == '是':
                        tt = '项目操作记录已更新,是否需要保存?'
                        xx = getvalue_yandn(tt)
                        saveact = xx[0]
                        if saveact == '是':
                            buildaction(infovalue1[0])
                    else:
                        tt = '是否需要记录项目操作以便下次抓取?'
                        xx = getvalue_yandn(tt)
                        saveact = xx[0]
                        if saveact == '是':
                            if infovalue1[0] != '' and infovalue1[0] != None:
                                buildaction(infovalue1[0])
                            else:
                                actioninfo = getvalue_saveact()
                                buildaction(actioninfo)
                if auto[0] == '是':
                    saveappend(allinfo)
                else:
                    saveways(allinfo)
                    if not judgeload == '是':
                        saveitem(infovalue1)
                enble2 = False
        tt = '是否需要再次抓取呢?'
        xx = getvalue_yandn(tt)
        restart = xx[0]
        if restart == '是':
            enble = True
        else:
            print('客官下次再来哟!')
            enble = False

#全自动化
def autoloadaction(eachname):
    enble = True
    getvalue = []
    judgeload = ''
    list = ['\\', '/', ':', '*', '?', '"', '<', '>', '|']
    while enble == True:
        try:
            ifload = '是'
            pastname = eachname
            for each in pastname:
                for un in list:
                    if un in each:
                        print('名字存在非法字符,请重新设置')
                        sys.exit()
            actpath = docpath + '\\' + pastname + '\\' + pastname + '.xlsx'
            if os.path.exists(actpath) == False:
                print('该项目不存在或未保存操作记录,请重新输入')
                sys.exit()
            else:
                openpath = openpyxl.load_workbook(actpath)
                sheet = openpath.get_sheet_names()
                ws = openpath.get_sheet_by_name(sheet[0])
                cx = ws.max_column
                value1 = []
                for i in range(0, 16):
                    x = ws.cell(row=i + 1, column=1).value
                    value1.append(x)
                getvalue.append(value1)
                for i in range(2, int(cx) + 1):
                    value2 = []
                    x1 = ws.cell(row=1, column=i).value
                    x2 = ws.cell(row=2, column=i).value
                    x3 = ws.cell(row=3, column=i).value
                    value2.append(x1)
                    value2.append(x2)
                    value2.append(x3)
                    getvalue.append(value2)
            judgeload = ifload + judgeload
            infovalue1[0] = pastname
            enble = False
        except:
            enble = True
    lens = len(getvalue)
    for i in range(0, lens):
        infovalue3.append([])
        infovalue3[i] = getvalue[i]
    return judgeload
def automainpro(eachname):
    print('开始更新', eachname)
    enble = True
    while enble == True:
        try:
            theheader = header
            allinfo = []
            auto = ['是']
            judgeload = autoloadaction(eachname)
            iflogin = ''
            openpath = openpyxl.load_workbook(doctxt)
            sheet = openpath.get_sheet_names()
            ws = openpath.get_sheet_by_name(sheet[1])
            rm = ws.max_row
            which = ['', '']
            for i in range(0, rm):
                val = ws.cell(row=i + 1, column=1).value
                if val == infovalue1[0]:
                    which[1] = infovalue3[0][0]
                    which[0] = ws.cell(row=i + 1, column=2).value
                    infovalue1[2] = ws.cell(row=i + 1, column=3).value
                    break
            if infovalue3[0][1] != None and infovalue3[0][1] != '':
                iflogin = '是' + iflogin
            if iflogin == '是':
                if not judgeload == '是':
                    print('可加载过去的项目自动登录')
                    judgeload = loadaction()
                theheader = postheader(judgeload)
                print('若是无法抓取到内容,请尝试重启程序重新登录')
            aimurl = which[1]
            aimurl = str(aimurl).replace(' ', '')
            if 'http' not in aimurl:
                aimurl = 'http://' + aimurl
            enble2 = True
            while enble2 == True:
                if which[0] == '1':
                    allinfo = getallinfo1(aimurl, theheader, auto)
                elif which[0] == '2':
                    allinfo = getallinfo2(aimurl, theheader, auto)
                elif which[0] == '3':
                    allinfo = getallinfo3(aimurl, auto)
                if allinfo == []:
                    print('什么也没有抓取到')
                    if iflogin == '是':
                        postcookie = []
                        cookie = savecookie(judgeload)
                        postcookie.append(cookie)
                        header['Cookie'] = postcookie[0]
                        enble2 = True
                    else:
                        enble2 = False
                else:
                    saveappend(allinfo)
                    enble2 = False
            print(eachname, '完成更新')
            enble = False
        except:
            enble = True
def allauto():
    autoname = []
    openpath = openpyxl.load_workbook(doctxt)
    sheet = openpath.get_sheet_names()
    ws1 = openpath.get_sheet_by_name(sheet[0])
    ws2 = openpath.get_sheet_by_name(sheet[1])
    rm = ws2.max_row
    ifauto = ws1.cell(row=3, column=1).value
    for i in range(0, rm):
        xxx = ws2.cell(row=i + 1, column=1).value
        if xxx != None:
            path = docpath + '\\' + xxx + '\\' + xxx + '.xlsx'
            if os.path.exists(path) != False:
                autoname.append(xxx)
    openpath.save(doctxt)
    if ifauto == '自动化':
        for each in autoname:
            automainpro(each)
        print('更新完毕,客官下次再来哟!')
    else:
        mainpro()

if __name__ == '__main__':
    try:
        allauto()
    except:
        print('客官还要再来哟!')